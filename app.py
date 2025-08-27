# Tambahkan impor yang diperlukan di bagian atas
from sqlalchemy import or_, func, and_  # Pastikan 'and_' juga diimport
from flask_login import login_required, current_user
from flask import request, render_template, redirect, url_for, flash
from werkzeug.middleware.proxy_fix import ProxyFix
from flask import Flask, render_template, url_for, request, redirect, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from flask_wtf.csrf import CSRFProtect
from pyngrok import ngrok
from flask import send_file, request
import qrcode
from flask import abort
import logging
import io
import string
import random
import requests
from flask_mail import Mail, Message
from flask import jsonify, request, session
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import openpyxl
from flask import make_response, request
from weasyprint import HTML
from decouple import config
from flask import session
import secrets
from datetime import date, datetime, time, timedelta
from sqlalchemy import func, or_  # Pastikan func dan or_ diimport
from sqlalchemy import or_
import math  # Tambahkan import math jika belum ada di bagian atas
# Impor untuk Login Manager & UserMixin
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
# Impor untuk Password Hashing
from werkzeug.utils import secure_filename  # Untuk mengamankan nama file
import uuid  # Untuk membuat nama file unik
import pytz
from werkzeug.security import generate_password_hash, check_password_hash
import os, base64

app = Flask(__name__)

app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
csrf = CSRFProtect(app)
wib_tz = pytz.timezone("Asia/Jakarta")
utc_tz = pytz.timezone("UTC")
# === Konfigurasi Upload Foto Profil ===
# Tentukan path absolut ke folder statis
basedir = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(basedir, 'static/profile_pics')
# Ekstensi file yang diizinkan
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# Buat folder jika belum ada
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# ======================================

limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"]
)
limiter.init_app(app)
# === Fungsi Helper untuk Cek Ekstensi ===


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def format_rupiah(value):
    """Memformat angka menjadi string format Rupiah (Rp xxx.xxx)."""
    try:
        # Konversi ke float dulu untuk jaga-jaga
        val = float(value)
        # Cek apakah angka punya desimal yang relevan (lebih dari .00)
        # Gunakan toleransi kecil untuk masalah floating point
        if abs(val - round(val)) < 0.0001:
            # Jika bisa dianggap integer, format tanpa desimal
            # Gunakan f-string untuk format dengan pemisah ribuan (,) lalu ganti ke (.)
            formatted_value = f"{round(val):,}".replace(",", ".")
        else:
            # Jika punya desimal, format dengan 2 desimal, lalu tukar pemisah
            # Format dulu ke US (,) ribuan (.) desimal
            us_formatted = "{:,.2f}".format(val)
            # Tukar: Koma jadi TEMP, Titik jadi Koma, TEMP jadi Titik
            id_formatted = us_formatted.replace(",", "TEMP_SEP").replace(
                ".", ",").replace("TEMP_SEP", ".")
            formatted_value = id_formatted

        return "Rp " + formatted_value
    except (ValueError, TypeError, OverflowError):
        # Jika input tidak valid (None, string non-angka, dll)
        return "Rp -"  # Tampilkan strip atau format default lain


app.jinja_env.filters['rupiah'] = format_rupiah
# ...existing code...
def short_rupiah(value, max_length=15):
    s = "{:,.0f}".format(value)
    if len(s) > max_length:
        return s[:max_length] + "..."
    return s

app.jinja_env.filters['short_rupiah'] = short_rupiah

# Konfigurasi Flask-Mail (isi sesuai SMTP Anda)
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'pramudyadimas510@gmail.com'
app.config['MAIL_PASSWORD'] = 'qmzhjtjpsljsozuo'
mail = Mail(app)

# ...existing code...

try:
    # Untuk Python 3.9+ (built-in)
    from zoneinfo import ZoneInfo
except ImportError:
    print("WARNING: zoneinfo module not found. Timezone features might be limited.")
    # Sediakan alternatif sederhana jika zoneinfo tidak ada

    class ZoneInfo:  # Dummy class
        def __init__(self, key): pass

# ... (import lain, app, db, model, dll) ...

# === Fungsi Custom Filter untuk Format WIB ===


def format_wib(utc_dt):
    """Mengkonversi datetime UTC ke string format WIB."""
    if not utc_dt:
        return "-"  # Tampilkan strip jika data tidak ada
    try:
        utc_tz = pytz.timezone("UTC")
        wib_tz = pytz.timezone("Asia/Jakarta")
        utc_dt_aware = utc_dt.replace(
            tzinfo=utc_tz) if not utc_dt.tzinfo else utc_dt
        wib_dt = utc_dt_aware.astimezone(wib_tz)
        return wib_dt.strftime('%d %b %Y, %H:%M:%S') + " WIB"
    except Exception as e:
        print(f"Error formatting date {utc_dt} to WIB: {e}")
        try:
            wib_dt_fallback = utc_dt + timedelta(hours=7)
            return wib_dt_fallback.strftime('%d %b %Y, %H:%M:%S') + " WIB (Fallback)"
        except:
            return str(utc_dt) + " UTC (Error)"
# === Akhir Fungsi Custom Filter ===


# Registrasikan filter (SETELAH app = Flask(__name__) dan SETELAH definisi fungsi format_wib)
app.jinja_env.filters['wib'] = format_wib

# --- Secret Key (PENTING!) ---
# Dibutuhkan Flask untuk mengamankan session (data login) dan flash messages.
# Ganti 'isi-dengan-kunci-rahasia-super-aman-dan-unik' dengan string acak yang panjang.
# Di aplikasi nyata, ini tidak boleh ditulis langsung di kode.
app.config['SECRET_KEY'] = os.environ.get(
    'SECRET_KEY', 'isi-dengan-kunci-rahasia-super-aman-dan-unik')
# --------------------------

# --- Konfigurasi Database ---
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///kantin.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
# --------------------------

# --- Konfigurasi Flask-Login ---
login_manager = LoginManager()
login_manager.init_app(app)
# Jika pengguna mencoba akses halaman yang butuh login tapi belum login,
# mereka akan diarahkan ke halaman yang namanya 'login' (nama fungsi route login kita).
login_manager.login_view = 'login'
# Pesan flash yang akan ditampilkan saat pengguna diarahkan
login_manager.login_message = "Anda harus login untuk mengakses halaman ini."
# Kategori pesan (untuk styling CSS nanti)
login_manager.login_message_category = "warning"
# --------------------------

# --- User Loader Callback ---
# Fungsi ini digunakan oleh Flask-Login untuk memuat user dari ID yang disimpan di session.


@login_manager.user_loader
def load_user(user_id):
    # Ambil user dari database berdasarkan ID
    return User.query.get(int(user_id))
# --------------------------

# --- Model Database (Update User) ---
# Tambahkan , UserMixin setelah db.Model


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(10), nullable=False)
    balance = db.Column(db.Float, nullable=False, default=0.0)
    profile_pic_filename = db.Column(db.String(100), nullable=True, default=None)
    kelas = db.Column(db.String(5), nullable=True)
    jurusan = db.Column(db.String(10), nullable=True)
    plain_password = db.Column(db.String(128)) 
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    active_border_id = db.Column(db.Integer, db.ForeignKey('border.id'), nullable=True)
    active_border = db.relationship('Border', backref=db.backref('users', lazy=True))
    pin_transfer = db.Column(db.String(6), nullable=True)
    credential_id = db.Column(db.String(255), nullable=True)  # Untuk WebAuthn
    webauthn_enabled = db.Column(db.Boolean, default=False)   # Status fitur WebAuthn
    email = db.Column(db.String(255), nullable=True)
    email_verified = db.Column(db.Boolean, default=False)

    # =====================================

    # Method set_password, check_password, __repr__ tetap sama
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f'<User {self.username} - {self.role}>'


# --- Class Transaction (TERPISAH dari User) ---
class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    related_user_id = db.Column(
        db.Integer, db.ForeignKey('user.id'), nullable=True)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    type = db.Column(db.String(20), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    user_balance_before = db.Column(db.Float, nullable=False)
    user_balance_after = db.Column(db.Float, nullable=False)
    description = db.Column(db.String(200), nullable=True)

    # Relasi
    user = db.relationship('User', foreign_keys=[user_id], backref=db.backref(
        'transactions', lazy='dynamic'))
    related_user = db.relationship('User', foreign_keys=[related_user_id])

    # ---> __repr__ untuk Transaction <---
    def __repr__(self):
        return f'<Transaction {self.id} ({self.type}) Rp {self.amount} for User {self.user_id}>'

    # Method untuk men-set password (otomatis hash)
    def set_password(self, password):
        # Membuat hash dari password yang diberikan
        self.password_hash = generate_password_hash(password)

    # Method untuk memeriksa apakah password yang diberikan cocok dengan hash
    def check_password(self, password):
        # Membandingkan password input dengan hash yang tersimpan
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f'<User {self.username} - {self.role}>'
# --------------------------


class TopUpRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(
        db.Integer, db.ForeignKey('user.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)  # Jumlah yg direquest
    request_timestamp = db.Column(
        db.DateTime, nullable=False, default=datetime.utcnow)  # Waktu request dibuat
    # Status: 'pending', 'approved', 'rejected'
    status = db.Column(db.String(20), nullable=False, default='pending')
    admin_id = db.Column(db.Integer, db.ForeignKey(
        'user.id'), nullable=True)  # ID Admin yg menyetujui/menolak
    # Waktu request diproses admin
    processed_timestamp = db.Column(db.DateTime, nullable=True)

    # Relasi untuk memudahkan akses data student/admin
    student = db.relationship('User', foreign_keys=[
                              student_id], backref=db.backref('topup_requests', lazy='dynamic'))
    admin = db.relationship('User', foreign_keys=[admin_id])
    bukti_transfer_filename = db.Column(db.String(100), nullable=True)

    def __repr__(self):
        # Representasi string objek (berguna untuk debugging)
        return f'<TopUpRequest {self.id} - Stud: {self.student_id} Amt: {self.amount} Stat: {self.status}>'


class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey(
        'user.id'), nullable=False, index=True)  # Penerima notif
    message = db.Column(db.String(255), nullable=False)  # Isi pesan notifikasi
    timestamp = db.Column(db.DateTime, nullable=False,
                          default=datetime.utcnow, index=True)  # Waktu notif dibuat
    # Status dibaca (default: belum)
    is_read = db.Column(db.Boolean, nullable=False, default=False)
    # Opsional: Link terkait notif (misal ke riwayat)
    related_link = db.Column(db.String(200), nullable=True)

    # Relasi ke User (agar mudah akses user dari notif)
    user = db.relationship('User', backref=db.backref(
        'notifications', lazy='dynamic', order_by='Notification.timestamp.desc()'))

    def __repr__(self):
        return f'<Notification {self.id} for User {self.user_id} - Read: {self.is_read}>'


# Tambahkan setelah class Notification

class Border(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    image_path = db.Column(db.String(200), nullable=False)
    required_transactions = db.Column(db.Integer, nullable=False)
    description = db.Column(db.String(200), nullable=True)
    jurusan = db.Column(db.String(50))  # Tambahkan kolom jurusan


class UserBorder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    border_id = db.Column(db.Integer, db.ForeignKey(
        'border.id'), nullable=False)
    unlocked_at = db.Column(db.DateTime, nullable=False,
                            default=datetime.utcnow)
    is_active = db.Column(db.Boolean, nullable=False, default=False)

    # Relasi
    user = db.relationship(
        'User', backref=db.backref('user_borders', lazy=True))
    border = db.relationship(
        'Border', backref=db.backref('user_borders', lazy=True))
    
class MarginSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nominal_per_margin = db.Column(db.Float, nullable=False, default=5000)  # contoh: 5000
    potongan_per_margin = db.Column(db.Float, nullable=False, default=500)  # contoh: 500


class Laporan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    bukti_filename = db.Column(db.String(200), nullable=True)
    keterangan = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    ditanggapi = db.Column(db.Boolean, default=False)


    user = db.relationship('User', backref=db.backref('laporan', lazy='dynamic'))

    def __repr__(self):
        return f'<Laporan {self.id} oleh User {self.user_id}>'

class AppConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    app_locked = db.Column(db.Boolean, nullable=False, default=False)
    lock_message = db.Column(db.String(255), nullable=True, default='Belum jam istirahat')
    locked_user_ids = db.Column(db.PickleType, nullable=True, default=[])
    locked_kelas = db.Column(db.PickleType, nullable=True, default=[])
    locked_jurusan = db.Column(db.PickleType, nullable=True, default=[])

def get_app_config():
    cfg = AppConfig.query.first()
    if not cfg:
        cfg = AppConfig(app_locked=False, lock_message='Belum jam istirahat')
        db.session.add(cfg)
        db.session.commit()
    return cfg

@app.context_processor
def inject_app_lock():
    cfg = None
    try:
        cfg = get_app_config()
    except Exception:
        cfg = None
    return {
        'app_locked': cfg.app_locked if cfg else False,
        'app_lock_message': cfg.lock_message if cfg else 'Layanan terkunci'
    }

# Admin view untuk toggle lock (hanya admin)
@app.route('/admin/app_lock', methods=['GET', 'POST'])
@login_required
def admin_app_lock():
    if current_user.role != 'admin':
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('dashboard'))
    cfg = get_app_config()

    # Ambil semua kelas & jurusan untuk dropdown
    all_kelas = sorted(set([u.kelas for u in User.query.filter(User.kelas.isnot(None)).all() if u.kelas]))
    all_jurusan = sorted(set([u.jurusan for u in User.query.filter(User.jurusan.isnot(None)).all() if u.jurusan]))

    if request.method == 'POST':
        action = request.form.get('action')
        # Kunci semua user (global)
        if action == 'lock':
            cfg.app_locked = True
            cfg.lock_message = request.form.get('lock_message') or cfg.lock_message
            db.session.commit()
            flash('Status kunci aplikasi diperbarui.', 'success')
            return redirect(url_for('admin_app_lock'))
        elif action == 'unlock':
            cfg.app_locked = False
            db.session.commit()
            flash('Status kunci aplikasi diperbarui.', 'success')
            return redirect(url_for('admin_app_lock'))
        # Kunci beberapa user
        elif action == 'update_user_lock':
            raw_ids = request.form.get('locked_user_ids', '')
            # Simpan sebagai list of int (atau str jika id bisa non-int)
            cfg.locked_user_ids = [i.strip() for i in raw_ids.split(',') if i.strip()]
            db.session.commit()
            flash('Daftar user yang dikunci diperbarui.', 'success')
            return redirect(url_for('admin_app_lock'))
        # Kunci kelas & jurusan
        elif action == 'update_kelas_jurusan_lock':
            cfg.locked_kelas = request.form.getlist('locked_kelas')
            cfg.locked_jurusan = request.form.getlist('locked_jurusan')
            print("DEBUG locked_kelas:", cfg.locked_kelas)
            print("DEBUG locked_jurusan:", cfg.locked_jurusan)
            db.session.add(cfg) 
            db.session.commit()
            flash('Daftar kelas & jurusan yang dikunci diperbarui.', 'success')
            return redirect(url_for('admin_app_lock'))

    # Pastikan field ada di config (untuk render awal)
    if not hasattr(cfg, 'locked_user_ids'):
        cfg.locked_user_ids = []
    if not hasattr(cfg, 'locked_kelas'):
        cfg.locked_kelas = []
    if not hasattr(cfg, 'locked_jurusan'):
        cfg.locked_jurusan = []

    return render_template(
        'admin/admin_app_lock.html',
        cfg=cfg,
        all_kelas=all_kelas,
        all_jurusan=all_jurusan,
        active_page='admin_app_lock'
    )

# helper check untuk route yang butuh block saat locked
def ensure_app_unlocked():
    cfg = get_app_config()
    # Global lock (semua user)
    if cfg.app_locked:
        return False, (cfg.lock_message or 'Aplikasi sedang dikunci oleh admin')

    # --- Lock khusus user, kelas, jurusan ---
    # Pastikan field ada di config, jika belum, set default
    locked_user_ids = getattr(cfg, 'locked_user_ids', []) or []
    locked_kelas = getattr(cfg, 'locked_kelas', []) or []
    locked_jurusan = getattr(cfg, 'locked_jurusan', []) or []

    # Jika user sudah login
    if current_user.is_authenticated:
        # Lock by user id
        if str(current_user.id) in [str(uid) for uid in locked_user_ids]:
            return False, cfg.lock_message or "Akun Anda sedang dikunci oleh admin."
        # Lock by kelas
        if current_user.kelas and current_user.kelas in locked_kelas:
            return False, cfg.lock_message or "Akses kelas Anda sedang dikunci oleh admin."
        # Lock by jurusan
        if current_user.jurusan and current_user.jurusan in locked_jurusan:
            return False, cfg.lock_message or "Akses jurusan Anda sedang dikunci oleh admin."

    return True, None


# --- Routes (Halaman Web) ---
# (Route index tetap sama)
@app.route('/')
def index():
    return render_template('login.html')

# (Route login akan diupdate nanti)

# --- Route Registrasi ---

@app.route('/send_email_otp', methods=['POST'])
@login_required
def send_email_otp():
    data = request.get_json()
    email = data.get('email')
    if not email:
        return jsonify({'success': False}), 400
    otp = ''.join(random.choices(string.digits, k=6))
    session['email_otp'] = otp
    session['email_otp_address'] = email
    # Kirim email
    msg = Message('Kode Verifikasi Email', sender=app.config['MAIL_USERNAME'], recipients=[email])
    msg.body = f'Kode verifikasi Anda: {otp}'
    try:
        mail.send(msg)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/verify_email_otp', methods=['POST'])
@login_required
def verify_email_otp():
    data = request.get_json()
    email = data.get('email')
    otp = data.get('otp')
    if session.get('email_otp') == otp and session.get('email_otp_address') == email:
        # Tandai email sudah diverifikasi di database user
        current_user.email = email
        current_user.email_verified = True
        db.session.commit()
        session.pop('email_otp', None)
        session.pop('email_otp_address', None)
        return jsonify({'success': True})
    return jsonify({'success': False})


# --- Route Login (Update) ---
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)


@app.errorhandler(400)
def bad_request(e):
    logger.error(f"Bad Request: {e}")
    flash("Permintaan tidak valid. Silakan coba lagi.", "danger")
    return redirect(request.url), 400


@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")  # Maksimal 5 percobaan login per menit per IP
def login():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard_content'))
        elif current_user.role == 'penjual':
            return redirect(url_for('penjual_dashboard'))
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        # --- Validasi reCAPTCHA ---
        recaptcha_response = request.form.get('g-recaptcha-response')
        if not recaptcha_response:
            flash('reCAPTCHA wajib diisi!', 'error')
            return redirect(url_for('login'))

        secret_key = "6LdJH7ErAAAAAIiNnKm3EVF7nWHikwBem0O7MSpl"
        payload = {
            'secret': secret_key,
            'response': recaptcha_response,
            'remoteip': request.remote_addr
        }
        r = requests.post('https://www.google.com/recaptcha/api/siteverify', data=payload)
        result = r.json()
        if not result.get('success'):
            flash('Verifikasi reCAPTCHA gagal, silakan coba lagi.', 'error')
            return redirect(url_for('login'))

        # --- Proses login seperti biasa ---
        user_id = request.form.get('id')
        password = request.form.get('password')
        remember = True if request.form.get('remember') else False

        user = User.query.get(user_id)
        if not user:
            flash('ID tidak ditemukan!', 'error')
            return redirect(url_for('login'))

        if not user.check_password(password):
            flash('Password salah!', 'error')
            return redirect(url_for('login'))

        login_user(user, remember=remember)

        if user.webauthn_enabled:
            return redirect(url_for('verify_device'))

        # Redirect berdasarkan role
        if user.role == 'admin':
            return redirect(url_for('admin_dashboard_content'))
        elif user.role == 'penjual':
            return redirect(url_for('penjual_dashboard'))
        else:
            return redirect(url_for('dashboard'))

    return render_template('login.html')


@app.route('/penjual/dashboard')
@login_required
def penjual_dashboard():
    if current_user.role != 'penjual':
        flash("Akses ditolak. Halaman ini hanya untuk penjual.", "danger")
        return redirect(url_for('dashboard'))

    # Zona Waktu WIB
    now_wib = datetime.now(wib_tz)
    today_wib = now_wib.date()
    today_wib_start_local = datetime.combine(today_wib, time.min).replace(tzinfo=wib_tz)
    tomorrow_wib_start_local = today_wib_start_local + timedelta(days=1)
    today_start_utc = today_wib_start_local.astimezone(utc_tz)
    tomorrow_start_utc = tomorrow_wib_start_local.astimezone(utc_tz)

    # Statistik Harian
    gross_earnings = 0.0  # Pendapatan kotor
    todays_transaction_count = 0
    total_transaksi = Transaction.query.filter_by(
        related_user_id=current_user.id, type='payment').count()
    try:
        sum_result = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.related_user_id == current_user.id,
            Transaction.type == 'payment',
            Transaction.timestamp >= today_start_utc,
            Transaction.timestamp < tomorrow_start_utc
        ).scalar()
        if sum_result is not None:
            gross_earnings = float(sum_result)

        todays_transaction_count = Transaction.query.filter(
            Transaction.related_user_id == current_user.id,
            Transaction.type == 'payment',
            Transaction.timestamp >= today_start_utc,
            Transaction.timestamp < tomorrow_start_utc
        ).count()
    except Exception as e:
        print(f"Error calculating vendor stats: {e}")
        flash("Gagal memuat statistik harian.", "warning")

    # Ambil margin setting dari database
    margin = MarginSetting.query.first()
    if not margin:
        margin_nominal = 5000
        margin_potongan = 500
    else:
        margin_nominal = margin.nominal_per_margin
        margin_potongan = margin.potongan_per_margin

    # Hitung total margin hari ini
    margin_count = int(gross_earnings // margin_nominal)
    margin_total = margin_count * margin_potongan

    # Pendapatan bersih
    total_earnings = gross_earnings - margin_total

    # Transaksi Terbaru (3 transaksi)
    recent_transactions = Transaction.query.options(
        db.joinedload(Transaction.user)
    ).filter(
        Transaction.related_user_id == current_user.id,
        Transaction.type == 'payment'
    ).order_by(Transaction.timestamp.desc()).limit(3).all()

    # Jumlah Notifikasi Belum Dibaca (opsional, jika ada model Notification)
    unread_count = 0
    try:
        unread_count = Notification.query.filter_by(
            user_id=current_user.id, is_read=False).count()
    except Exception as e:
        print(f"Error fetching notifications for vendor {current_user.id}: {e}")

    # Data untuk Grafik (7 hari terakhir)
    chart_labels = []
    chart_data = []
    for i in range(6, -1, -1):
        day = today_wib - timedelta(days=i)
        chart_labels.append(day.strftime('%d %b'))
        start = datetime.combine(day, time.min).replace(tzinfo=wib_tz).astimezone(utc_tz)
        end = datetime.combine(day, time.max).replace(tzinfo=wib_tz).astimezone(utc_tz)
        sum_result = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.related_user_id == current_user.id,
            Transaction.type == 'payment',
            Transaction.timestamp >= start,
            Transaction.timestamp <= end
        ).scalar() or 0
        chart_data.append(float(sum_result))

    # Data untuk template
    data = {
        'qr_code_url': url_for('my_qr_code'),
        'balance': current_user.balance,
        'todays_earnings': total_earnings,           # bersih
        'gross_earnings': gross_earnings,            # kotor
        'margin_total': margin_total,                # margin
        'todays_transaction_count': todays_transaction_count,
        'recent_transactions': recent_transactions,
        'unread_count': unread_count,
        'is_active': current_user.is_active,
        'chart_labels': chart_labels,
        'chart_data': chart_data
    }

    return render_template('penjual_dashboard.html', **data, total_transaksi=total_transaksi)


@app.route('/penjual/history')
@login_required
def history_vendor():
    if current_user.role != 'penjual':
        flash("Akses ditolak. Halaman ini hanya untuk penjual.", "danger")
        return redirect(url_for('dashboard'))

    # Parameter filter & sorting
    page = request.args.get('page', 1, type=int)
    per_page = 10
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    sort = request.args.get('sort', 'desc')
    transaction_id = request.args.get('transaction_id', type=int)  # Tambah ini

    # Zona Waktu
    now_wib = datetime.now(wib_tz)
    today_wib = now_wib.date()
    default_start_dt = today_wib - timedelta(days=29)
    default_end_dt = today_wib

    # Parsing tanggal
    try:
        start_dt = date.fromisoformat(
            start_date_str) if start_date_str else default_start_dt
    except ValueError:
        start_dt = default_start_dt
    try:
        end_dt = date.fromisoformat(
            end_date_str) if end_date_str else default_end_dt
    except ValueError:
        end_dt = default_end_dt
    if start_dt > end_dt:
        start_dt, end_dt = default_start_dt, default_end_dt

    # Konversi ke UTC
    start_wib_day_local = datetime.combine(
        start_dt, time.min).replace(tzinfo=wib_tz)
    end_wib_day_local = datetime.combine(
        end_dt, time.max).replace(tzinfo=wib_tz)
    start_utc = start_wib_day_local.astimezone(utc_tz)
    end_utc = end_wib_day_local.astimezone(utc_tz)

    payment_query = Transaction.query.options(
        db.joinedload(Transaction.user)
    ).filter(
        Transaction.related_user_id == current_user.id,
        Transaction.type == 'payment',
        Transaction.timestamp >= start_utc,
        Transaction.timestamp <= end_utc
    )

    withdrawal_query = Transaction.query.options(
        db.joinedload(Transaction.user)
    ).filter(
        Transaction.user_id == current_user.id,
        Transaction.type.in_(['withdrawal', 'cashout_vendor']),
        Transaction.timestamp >= start_utc,
        Transaction.timestamp <= end_utc
    )

    # Gabungkan kedua query
    all_query = payment_query.union_all(withdrawal_query)

    # Filter berdasarkan ID transaksi jika diisi
    if transaction_id:
        all_query = all_query.filter(Transaction.id == transaction_id)

    # Sorting
    if sort == 'asc':
        all_query = all_query.order_by(Transaction.timestamp.asc())
    elif sort == 'max':
        all_query = all_query.order_by(Transaction.amount.desc())
    elif sort == 'min':
        all_query = all_query.order_by(Transaction.amount.asc())
    else:  # default 'desc'
        all_query = all_query.order_by(Transaction.timestamp.desc())

    # Pagination
    transactions_pagination = all_query.paginate(page=page, per_page=per_page)
    total_transaksi = all_query.count()

    return render_template(
        'penjual_history.html',
        transactions=transactions_pagination.items,
        pagination=transactions_pagination,
        filter_start_date=start_dt.strftime('%Y-%m-%d'),
        filter_end_date=end_dt.strftime('%Y-%m-%d'),
        total_transaksi=total_transaksi,
        filter_transaction_id=transaction_id  # Kirim ke template
    )


@app.route('/penjual/profile')
@login_required
def penjual_profile():
    if current_user.role != 'penjual':
        flash('Akses tidak diizinkan.', 'danger')
        return redirect(url_for('penjual_dashboard'))
    return render_template('penjual_profile.html')

# --- Route Logout ---


@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.pop('device_verified', None)
    flash('Anda telah berhasil logout.', 'info')
    return redirect(url_for('index'))


# --- Route Dashboard ---
# Pastikan SEMUA import ini ada di bagian atas app.py
try:
    from zoneinfo import ZoneInfo
except ImportError:
    # Fallback sederhana jika zoneinfo tidak ada
    from datetime import timezone, timedelta as ZoneInfo
    print("Warning: zoneinfo module not found, using simple timedelta fallback for WIB.")
    wib = None  # Tandai bahwa zoneinfo tidak tersedia
else:
    try:
        wib = ZoneInfo("Asia/Jakarta")
    except Exception as e:
        print(f"Warning: Could not load Asia/Jakarta timezone: {e}")
        wib = None


# ... (kode app = Flask(), config, filter, login_manager, model lain, route lain) ...


@app.route('/dashboard')
@login_required
def dashboard():
    # Jika user mengaktifkan WebAuthn dan belum verifikasi device, redirect ke /verify_device
    if current_user.webauthn_enabled and not session.get('device_verified'):
        return redirect(url_for('verify_device'))
    if current_user.role != 'siswa':
        return redirect(url_for('admin_dashboard_content'))

    # Tambahkan ini sebelum query
    today = datetime.utcnow().date()

    # Get recent transactions
    recent_transactions = (Transaction.query
        .filter(or_(
            Transaction.user_id == current_user.id,
            Transaction.related_user_id == current_user.id
        ))
        .order_by(Transaction.timestamp.desc())
        .limit(5)
        .all())

    # Calculate today's spending
    todays_spending = (Transaction.query
        .filter(
            Transaction.user_id == current_user.id,
            Transaction.type.in_(['payment', 'transfer']),
            func.date(Transaction.timestamp) == today
        )
        .with_entities(func.sum(Transaction.amount))
        .scalar() or 0)


    # Get notifications
    recent_notifications = (Notification.query
    .filter(Notification.user_id == current_user.id)
    .order_by(Notification.timestamp.desc())
    .all())

    # Filter border notifications
    filtered_notifications = []
    for notif in recent_notifications:
        if 'border' not in notif.message.lower():
            filtered_notifications.append(notif)
        elif current_user.jurusan and current_user.jurusan.lower() in notif.message.lower():
            filtered_notifications.append(notif)

        filtered_notifications = filtered_notifications[:3]

    # Count unread notifications
    unread_count = (Notification.query
        .filter(
            Notification.user_id == current_user.id,
            Notification.is_read == False
        )
        .count())

    return render_template('siswa/dashboard_siswa.html',
                        recent_transactions=recent_transactions,
                        todays_spending=todays_spending,
                        unread_count=unread_count,
                        recent_notifications=filtered_notifications,
                        active_page='dashboard',)


@app.route('/admin/topup', methods=['GET', 'POST'])
@login_required
def admin_topup():
    if current_user.role != 'admin':
        flash("Akses ditolak. Fitur ini hanya untuk admin.", "danger")
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        user_id_to_topup = request.form.get('user_id')
        amount_str = request.form.get('amount')

        student = User.query.get(user_id_to_topup)
        amount = 0.0
        error_message = None

        if not student or student.role != 'siswa':
            error_message = "Siswa tidak ditemukan atau ID tidak valid."
        elif not amount_str:
            error_message = "Jumlah top up wajib diisi."
        else:
            try:
                amount = float(amount_str)
                if amount <= 0:
                    error_message = "Jumlah top up harus lebih dari 0."
            except ValueError:
                error_message = "Jumlah top up harus berupa angka."

        if error_message:
            flash(error_message, "danger")
            return render_template('admin/admin_topup.html', active_page='admin_topup')

        try:
            balance_before = student.balance
            student.balance += amount
            balance_after = student.balance

            new_transaction = Transaction(
                user_id=current_user.id,
                related_user_id=student.id,
                type='topup',
                amount=amount,
                user_balance_before=balance_before,
                user_balance_after=balance_after,
                description=f"Top up saldo untuk {student.username} oleh Admin"
            )
            db.session.add(new_transaction)
            db.session.commit()

            flash(
                f"Top up untuk {student.username} sebesar {format_rupiah(amount)} berhasil. Saldo baru: {format_rupiah(student.balance)}", "success")

        except Exception as e:
            db.session.rollback()
            flash(f"Terjadi kesalahan saat proses top up: {str(e)}", "danger")

        return redirect(url_for('admin_topup'))

    # GET: Tampilkan form tanpa daftar siswa
    return render_template('admin/admin_topup.html', active_page='admin_topup')

@app.route('/admin/get_student_username')
@login_required
def get_student_username():
    if current_user.role != 'admin':
        return {'error': 'unauthorized'}, 403
    user_id = request.args.get('id')
    user = User.query.get(user_id)
    if user and user.role == 'siswa':
        return {'username': user.username}
    return {'username': None}

# --- Route untuk Admin Dashboard (Opsional tapi bagus) ---


@app.route('/admin/dashboard')
@login_required
def admin_dashboard_content():
    if current_user.role != 'admin':
        flash("Akses ditolak.", "danger")
        return redirect(url_for('dashboard'))

    # Statistik pengguna
    total_admins = User.query.filter_by(role='admin').count()
    total_vendors = User.query.filter_by(role='penjual').count()
    total_students = User.query.filter_by(role='siswa').count()
    active_vendors = User.query.filter_by(role='penjual', is_active=True).count()
    active_students = User.query.filter_by(role='siswa', is_active=True).count()

    # Statistik saldo
    total_student_balance = db.session.query(db.func.coalesce(db.func.sum(User.balance), 0)).filter_by(role='siswa', is_active=True).scalar()
    total_vendor_balance = db.session.query(db.func.coalesce(db.func.sum(User.balance), 0)).filter_by(role='penjual', is_active=True).scalar()

    # Statistik aktivitas hari ini
    today = datetime.utcnow().date()
    start_today = datetime.combine(today, datetime.min.time())
    end_today = datetime.combine(today, datetime.max.time())

    total_topups_today = db.session.query(db.func.coalesce(db.func.sum(Transaction.amount), 0))\
        .filter(Transaction.type == 'topup', Transaction.timestamp >= start_today, Transaction.timestamp <= end_today).scalar()
    total_payments_today = db.session.query(db.func.coalesce(db.func.sum(Transaction.amount), 0))\
        .filter(Transaction.type == 'payment', Transaction.timestamp >= start_today, Transaction.timestamp <= end_today).scalar()
    total_cashouts_today = db.session.query(db.func.coalesce(db.func.sum(Transaction.amount), 0))\
        .filter(Transaction.type == 'cashout_vendor', Transaction.timestamp >= start_today, Transaction.timestamp <= end_today).scalar()

    # Statistik tambahan
    total_transactions_today = Transaction.query.filter(
        Transaction.timestamp >= start_today,
        Transaction.timestamp <= end_today
    ).count()

    pending_topup_requests = TopUpRequest.query.filter_by(status='pending').count()

    total_transfer_today = db.session.query(db.func.coalesce(db.func.sum(Transaction.amount), 0)).filter(
        Transaction.type == 'transfer',
        Transaction.timestamp >= start_today,
        Transaction.timestamp <= end_today
    ).scalar()

    students_no_topup = User.query.filter_by(role='siswa', is_active=True, balance=0).count()

    students_low_balance = User.query.filter(
        User.role == 'siswa',
        User.is_active == True,
        User.balance < 10000
    ).count()

    vendors_today = db.session.query(User.username, db.func.count(Transaction.id)).join(Transaction, Transaction.related_user_id == User.id).filter(
        User.role == 'penjual',
        Transaction.type == 'payment',
        Transaction.timestamp >= start_today,
        Transaction.timestamp <= end_today
    ).group_by(User.username).order_by(db.func.count(Transaction.id).desc()).first()

    students_today = db.session.query(User.username, db.func.count(Transaction.id)).join(Transaction, Transaction.user_id == User.id).filter(
        User.role == 'siswa',
        Transaction.timestamp >= start_today,
        Transaction.timestamp <= end_today
    ).group_by(User.username).order_by(db.func.count(Transaction.id).desc()).first()

    laporan_pending = Laporan.query.filter_by(ditanggapi=False).count()

    nonactive_students = User.query.filter_by(role='siswa', is_active=False).count()
    nonactive_vendors = User.query.filter_by(role='penjual', is_active=False).count()

    notifikasi_today = Notification.query.filter(
        Notification.timestamp >= start_today,
        Notification.timestamp <= end_today
    ).count()

    stats = {
        'total_admins': total_admins,
        'total_vendors': total_vendors,
        'total_students': total_students,
        'active_vendors': active_vendors,
        'active_students': active_students,
        'total_student_balance': total_student_balance,
        'total_vendor_balance': total_vendor_balance,
        'total_topups_today': total_topups_today,
        'total_payments_today': total_payments_today,
        'total_cashouts_today': total_cashouts_today,
        'total_transactions_today': total_transactions_today,
        'pending_topup_requests': pending_topup_requests,
        'total_transfer_today': total_transfer_today,
        'students_no_topup': students_no_topup,
        'students_low_balance': students_low_balance,
        'vendors_today': vendors_today[0] if vendors_today else '-',
        'vendors_today_count': vendors_today[1] if vendors_today else 0,
        'students_today': students_today[0] if students_today else '-',
        'students_today_count': students_today[1] if students_today else 0,
        'laporan_pending': laporan_pending,
        'nonactive_students': nonactive_students,
        'nonactive_vendors': nonactive_vendors,
        'notifikasi_today': notifikasi_today
    }

    # Grafik aktivitas transaksi 7 hari terakhir
    today = datetime.utcnow().date()
    last_7_days = [today - timedelta(days=i) for i in range(6, -1, -1)]
    chart_labels = []
    chart_topup = []
    chart_payment = []
    chart_cashout = []
    chart_transactions = []
    chart_transfers = []
    chart_students_no_topup = []
    chart_students_low_balance = []
    chart_active_vendors = []
    chart_active_students = []

    for d in last_7_days:
        start = datetime.combine(d, datetime.min.time())
        end = datetime.combine(d, datetime.max.time())
        chart_labels.append(d.strftime('%d %b'))
        chart_topup.append(Transaction.query.filter(Transaction.type == 'topup', Transaction.timestamp >= start, Transaction.timestamp <= end).count())
        chart_payment.append(Transaction.query.filter(Transaction.type == 'payment', Transaction.timestamp >= start, Transaction.timestamp <= end).count())
        chart_cashout.append(Transaction.query.filter(Transaction.type == 'cashout_vendor', Transaction.timestamp >= start, Transaction.timestamp <= end).count())
        chart_transactions.append(Transaction.query.filter(Transaction.timestamp >= start, Transaction.timestamp <= end).count())
        chart_transfers.append(Transaction.query.filter(Transaction.type == 'transfer', Transaction.timestamp >= start, Transaction.timestamp <= end).count())
        chart_students_no_topup.append(User.query.filter_by(role='siswa', is_active=True, balance=0).count())
        chart_students_low_balance.append(User.query.filter(User.role == 'siswa', User.is_active == True, User.balance < 10000).count())
        # Penjual teraktif dan siswa teraktif bisa diisi dengan angka dummy jika ingin chart, misal 0
        chart_active_vendors.append(0)
        chart_active_students.append(0)

    return render_template(
        '/admin/dashboard_admin_content.html',
        stats=stats,
        chart_labels=chart_labels,
        chart_topup=chart_topup,
        chart_payment=chart_payment,
        chart_cashout=chart_cashout,
        chart_transactions=chart_transactions,
        chart_transfers=chart_transfers,
        chart_students_no_topup=chart_students_no_topup,
        chart_students_low_balance=chart_students_low_balance,
        chart_active_vendors=chart_active_vendors,
        chart_active_students=chart_active_students,
        now=datetime.utcnow(),
        active_page='admin_dashboard_content'
    )


@app.route('/pay', methods=['POST'])
@login_required  # Harus login untuk bayar
def pay():
    # 1. Otorisasi: Pastikan yang bayar adalah siswa
    if current_user.role != 'siswa':
        flash("Hanya siswa yang dapat melakukan pembayaran.", "danger")
        return redirect(url_for('dashboard'))

    # CEK APP LOCK DI AWAL
    ok, msg = ensure_app_unlocked()
    if not ok:
        flash(msg, 'warning')
        return redirect(url_for('dashboard'))

    # 2. Ambil data dari form yang dikirim
    vendor_id = request.form.get('vendor_id')
    amount_str = request.form.get('amount')

    # 3. Validasi Input
    vendor = User.query.get(vendor_id)  # Cari penjual berdasarkan ID
    amount = 0.0
    error_message = None  # Variabel untuk menampung pesan error

    if not vendor or vendor.role != 'penjual':
        error_message = "Penjual yang dipilih tidak valid."
    elif not amount_str:
        error_message = "Jumlah pembayaran wajib diisi."
    else:
        try:
            amount = float(amount_str)
            if amount <= 0:  # Harus positif
                error_message = "Jumlah pembayaran harus lebih dari 0."
        except ValueError:
            error_message = "Jumlah pembayaran harus berupa angka."

    # 4. Validasi Saldo: Cek apakah saldo siswa mencukupi
    # Lakukan ini HANYA jika validasi sebelumnya lolos (error_message masih None)
    if not error_message:
        if current_user.balance < amount:
            error_message = f"Saldo Anda tidak mencukupi! Saldo saat ini: Rp {'%.2f' % current_user.balance}"

    # Jika ada error dari validasi di atas, tampilkan pesan dan kembali ke dashboard
    if error_message:
        flash(error_message, "danger")
        return redirect(url_for('dashboard'))

    # 5. Proses Transaksi Database (Jika semua validasi lolos)
    try:
        # Catat saldo sebelum transfer
        student_balance_before = current_user.balance
        vendor_balance_before = vendor.balance  # Saldo penjual sebelum

        # Lakukan transfer saldo
        current_user.balance -= amount
        vendor.balance += amount

        # Catat saldo siswa sesudah transfer
        student_balance_after = current_user.balance

        # 6. Buat record transaksi pembayaran
        payment_transaction = Transaction(
            user_id=current_user.id,         # ID Siswa yang membayar
            related_user_id=vendor.id,       # ID Penjual yang menerima
            type='payment',                  # Jenis transaksi
            amount=amount,                   # Jumlah yang dibayar
            user_balance_before=student_balance_before,
            user_balance_after=student_balance_after,
            # Deskripsi
            description=f"Pembayaran ke penjual: {vendor.username}"
        )
        db.session.add(payment_transaction)  # Tambahkan transaksi ke sesi

        # (Opsional) Buat juga record transaksi dari sisi Penjual jika perlu detail di history penjual
        # vendor_transaction = Transaction(...) # type='penerimaan' atau sejenisnya
        # db.session.add(vendor_transaction)

        # Commit semua perubahan ke database (saldo siswa, saldo penjual, transaksi baru)
        # Ini penting dilakukan dalam satu blok try-commit agar atomik
        db.session.commit()

        flash(
            f"Pembayaran sebesar Rp {'%.2f' % amount} kepada {vendor.username} BERHASIL!", "success")

    except Exception as e:
        # Jika terjadi error saat commit, batalkan semua perubahan di sesi ini
        db.session.rollback()
        flash(
            f"GAGAL melakukan pembayaran: Terjadi kesalahan sistem. ({str(e)})", "danger")

    # Redirect kembali ke dashboard setelah selesai (sukses atau gagal)
    return redirect(url_for('dashboard'))


@app.route('/cashout', methods=['POST'])
@login_required
def cash_out():
    if current_user.role != 'penjual':
        flash("Hanya penjual yang dapat melakukan penarikan saldo.", "danger")
        return redirect(url_for('dashboard'))

    try:
        amount_to_cash_out = float(request.form.get('amount_to_cash_out', 0))
        if amount_to_cash_out < 5000:
            flash("Minimal penarikan adalah Rp5.000.", "danger")
            return redirect(url_for('penjual_dashboard'))
        if amount_to_cash_out > current_user.balance:
            flash("Saldo tidak mencukupi untuk penarikan.", "danger")
            return redirect(url_for('penjual_dashboard'))

        # ...ambil margin dari database seperti sebelumnya...
        margin = MarginSetting.query.first()
        if not margin:
            margin_nominal = 5000
            margin_potongan = 500
        else:
            margin_nominal = margin.nominal_per_margin
            margin_potongan = margin.potongan_per_margin

        margin_count = int(amount_to_cash_out // margin_nominal)
        total_potongan = margin_count * margin_potongan
        bersih_diterima = amount_to_cash_out - total_potongan

        vendor_balance_before = current_user.balance
        admin_bank_mini = User.query.filter_by(role='admin').first()
        admin_balance_before = admin_bank_mini.balance

        current_user.balance -= amount_to_cash_out
        admin_bank_mini.balance += bersih_diterima

        cashout_transaction = Transaction(
            user_id=current_user.id,
            related_user_id=admin_bank_mini.id,
            type='cashout_vendor',
            amount=bersih_diterima,
            user_balance_before=vendor_balance_before,
            user_balance_after=current_user.balance,
            description=f"Penarikan tunai oleh {current_user.username} (potongan {total_potongan})"
        )
        db.session.add(cashout_transaction)
        db.session.commit()

        flash(
            f"Penarikan tunai sebesar {format_rupiah(amount_to_cash_out)} berhasil diproses! Potongan: {format_rupiah(total_potongan)}. Saldo diterima: {format_rupiah(bersih_diterima)}", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Gagal melakukan penarikan tunai: {str(e)}", "danger")

    return redirect(url_for('penjual_dashboard'))


@app.route('/my_qr_code.png')
@login_required
def my_qr_code():
    if current_user.role != 'penjual':
        return "Akses ditolak", 403

    background_path = os.path.join('static', 'qr_bg', 'background_a5.png')
    try:
        canvas = Image.open(background_path).convert("RGBA")
    except Exception as e:
        canvas = Image.new("RGBA", (1748, 2480), (255, 255, 255, 255))

    nama_penjual = current_user.username
    draw = ImageDraw.Draw(canvas)
    try:
        font = ImageFont.truetype("arial.ttf", 80)
    except:
        font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), nama_penjual, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    text_x = (1748 - text_width) // 2
    text_y = 600

    draw.text((text_x, text_y), nama_penjual, fill=(0, 0, 0, 255), font=font)

        # Tambahkan ID penjual di bawah nama penjual
    id_penjual = f"ID: {current_user.id}"
    try:
        font_id = ImageFont.truetype("arial.ttf", 50)
    except:
        font_id = ImageFont.load_default()
    bbox_id = draw.textbbox((0, 0), id_penjual, font=font_id)
    id_width = bbox_id[2] - bbox_id[0]
    id_height = bbox_id[3] - bbox_id[1]
    id_x = (1748 - id_width) // 2
    id_y = text_y + text_height + 50  # 10px di bawah nama penjual

    draw.text((id_x, id_y), id_penjual, fill=(0, 0, 0, 255), font=font_id)

    data_to_encode = str(current_user.id)
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=1,
    )
    qr.add_data(data_to_encode)
    qr.make(fit=True)
    # QR code dengan background putih
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")

    qr_size = 800
    qr_img = qr_img.resize((qr_size, qr_size))
    qr_x = (1748 - qr_size) // 2
    qr_y = 900  # bebas atur

    # Tempel QR code dengan mask agar putihnya tetap
    canvas.paste(qr_img, (qr_x, qr_y), qr_img)

    img_io = io.BytesIO()
    canvas.save(img_io, 'PNG')
    img_io.seek(0)
    return send_file(img_io, mimetype='image/png')

@app.route('/my_qr_code_only.png')
@login_required
def my_qr_code_only():
    if current_user.role != 'penjual':
        return "Akses ditolak", 403

    data_to_encode = str(current_user.id)
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=1,
    )
    qr.add_data(data_to_encode)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")

    qr_size = 800
    qr_img = qr_img.resize((qr_size, qr_size))

    img_io = io.BytesIO()
    qr_img.save(img_io, 'PNG')
    img_io.seek(0)
    return send_file(img_io, mimetype='image/png')


@app.context_processor
def inject_now():
    return {'now': datetime.utcnow()}


@app.route('/admin/users')
@login_required
def admin_users():
    if current_user.role != 'admin':
        flash("Akses ditolak. Halaman ini hanya untuk admin.", "danger")
        return redirect(url_for('dashboard'))

    # Ambil parameter filter dari URL
    page = request.args.get('page', 1, type=int)
    per_page = 12
    filter_kelas = request.args.get('kelas')
    filter_jurusan = request.args.get('jurusan')
    filter_user_id = request.args.get('user_id', type=int)  # Ganti dari username

    query = User.query

    # Filter berdasarkan kelas
    if filter_kelas:
        query = query.filter(User.kelas == filter_kelas)

    # Filter berdasarkan jurusan
    if filter_jurusan:
        query = query.filter(User.jurusan == filter_jurusan)

    # Filter berdasarkan ID user
    if filter_user_id:
        query = query.filter(User.id == filter_user_id)

    users_pagination = query.order_by(
        User.id.asc()).paginate(page=page, per_page=per_page)

    return render_template(
        '/admin/admin_users.html',
        users=users_pagination.items,
        pagination=users_pagination,
        filter_kelas=filter_kelas,
        filter_jurusan=filter_jurusan,
        filter_user_id=filter_user_id,  # Ganti dari filter_username
        active_page='admin_users')


@app.route('/admin/transactions')
@login_required
def admin_transactions():
    if current_user.role != 'admin':
        flash("Akses ditolak. Halaman ini hanya untuk admin.", "danger")
        return redirect(url_for('dashboard'))

    # Ambil parameter filter dari URL
    page = request.args.get('page', 1, type=int)
    per_page = 8
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    filter_type = request.args.get('filter_type', 'all')
    username = request.args.get('username')
    transaction_id = request.args.get('transaction_id', type=int)  # Tambah ini

    query = Transaction.query

    # Filter berdasarkan ID transaksi jika diisi
    if transaction_id:
        query = query.filter(Transaction.id == transaction_id)
    else:
        # Filter berdasarkan tanggal
        if start_date:
            query = query.filter(Transaction.timestamp >= start_date)
        if end_date:
            query = query.filter(Transaction.timestamp <= end_date)
        # Filter berdasarkan tipe transaksi
        if filter_type != 'all':
            query = query.filter(Transaction.type == filter_type)
        # Filter berdasarkan username
        if username:
            query = query.join(User, Transaction.user_id == User.id).filter(
                User.username.ilike(f"%{username}%"))

    transactions_pagination = query.order_by(
        Transaction.timestamp.desc()).paginate(page=page, per_page=per_page)

    return render_template(
        '/admin/admin_transactions.html',
        transactions=transactions_pagination.items,
        pagination=transactions_pagination,
        filter_start_date=start_date,
        filter_end_date=end_date,
        selected_type_filter=filter_type,
        filter_username=username,
        filter_transaction_id=transaction_id,  # Kirim ke template
        active_page='admin_transactions')


@app.route('/print/daily_summary')
@login_required
def print_daily_summary():
    # 1. Otorisasi: Pastikan yang akses adalah penjual
    if current_user.role != 'penjual':
        flash("Hanya penjual yang dapat mengakses fitur ini.", "danger")
        return redirect(url_for('dashboard'))

    # 2. Dapatkan Zona Waktu & Tanggal Hari Ini (WIB)
    try:
        wib = ZoneInfo("Asia/Jakarta")
    except Exception:
        wib = None
    now_wib = datetime.now(wib) if wib else datetime.utcnow() + timedelta(hours=7)
    today_wib = now_wib.date()

    # 3. Tentukan Batas Waktu UTC untuk Query Hari Ini (WIB)
    today_wib_start_local = datetime.combine(today_wib, time.min)
    if wib:
        today_wib_start_local = today_wib_start_local.replace(tzinfo=wib)
    tomorrow_wib_start_local = today_wib_start_local + timedelta(days=1)
    today_start_utc_for_calc = today_wib_start_local.astimezone(ZoneInfo("UTC")) if wib and isinstance(wib, ZoneInfo) else today_wib_start_local - timedelta(hours=7)
    tomorrow_start_utc_for_calc = tomorrow_wib_start_local.astimezone(ZoneInfo("UTC")) if wib and isinstance(wib, ZoneInfo) else tomorrow_wib_start_local - timedelta(hours=7)

    # 4. Hitung Total Pendapatan Kotor Hari Ini
    gross_earnings = 0.0
    try:
        sum_result = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.related_user_id == current_user.id,
            Transaction.type == 'payment',
            Transaction.timestamp >= today_start_utc_for_calc,
            Transaction.timestamp < tomorrow_start_utc_for_calc
        ).scalar()
        if sum_result is not None:
            gross_earnings = float(sum_result)
    except Exception as e:
        print(f"Error menghitung pendapatan harian untuk print: {e}")

    # 5. Hitung Margin Hari Ini
    margin = MarginSetting.query.first()
    if not margin:
        margin_nominal = 5000
        margin_potongan = 500
    else:
        margin_nominal = margin.nominal_per_margin
        margin_potongan = margin.potongan_per_margin

    margin_count = int(gross_earnings // margin_nominal)
    margin_total = margin_count * margin_potongan

    # 6. Hitung Pendapatan Bersih
    total_earnings = gross_earnings - margin_total

    # 7. Ambil Daftar Transaksi Pembayaran Diterima Hari Ini
    daily_transactions = []
    try:
        daily_transactions = Transaction.query.options(
            db.joinedload(Transaction.user)
        ).filter(
            Transaction.related_user_id == current_user.id,
            Transaction.type == 'payment',
            Transaction.timestamp >= today_start_utc_for_calc,
            Transaction.timestamp < tomorrow_start_utc_for_calc
        ).order_by(Transaction.timestamp.asc()).all()
    except Exception as e:
        print(f"Error mengambil transaksi harian untuk print: {e}")
        flash("Gagal mengambil rincian transaksi untuk dicetak.", "warning")

    # 8. Siapkan data untuk dikirim ke template cetak
    report_data = {
        "vendor_name": current_user.username,
        "report_date_str": today_wib.strftime('%A, %d %B %Y'),
        "total_earnings": total_earnings,      # bersih
        "gross_earnings": gross_earnings,      # kotor
        "margin_total": margin_total,          # margin
        "transactions": daily_transactions,
        "print_time_wib": format_wib(datetime.utcnow())
    }

    # 9. Render template khusus untuk cetak
    return render_template('vendor_daily_print.html', **report_data)

@app.route('/print/daily_summary/pdf')
@login_required
def print_daily_summary_pdf():
    # 1. Otorisasi: Pastikan yang akses adalah penjual
    if current_user.role != 'penjual':
        flash("Hanya penjual yang dapat mengakses fitur ini.", "danger")
        return redirect(url_for('dashboard'))

    # 2. Dapatkan Zona Waktu & Tanggal Hari Ini (WIB)
    try:
        wib = ZoneInfo("Asia/Jakarta")
    except Exception:
        wib = None
    now_wib = datetime.now(wib) if wib else datetime.utcnow() + timedelta(hours=7)
    today_wib = now_wib.date()

    # 3. Tentukan Batas Waktu UTC untuk Query Hari Ini (WIB)
    today_wib_start_local = datetime.combine(today_wib, time.min)
    if wib:
        today_wib_start_local = today_wib_start_local.replace(tzinfo=wib)
    tomorrow_wib_start_local = today_wib_start_local + timedelta(days=1)
    today_start_utc_for_calc = today_wib_start_local.astimezone(ZoneInfo("UTC")) if wib and isinstance(wib, ZoneInfo) else today_wib_start_local - timedelta(hours=7)
    tomorrow_start_utc_for_calc = tomorrow_wib_start_local.astimezone(ZoneInfo("UTC")) if wib and isinstance(wib, ZoneInfo) else tomorrow_wib_start_local - timedelta(hours=7)

    # 4. Hitung Total Pendapatan Kotor Hari Ini
    gross_earnings = 0.0
    try:
        sum_result = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.related_user_id == current_user.id,
            Transaction.type == 'payment',
            Transaction.timestamp >= today_start_utc_for_calc,
            Transaction.timestamp < tomorrow_start_utc_for_calc
        ).scalar()
        if sum_result is not None:
            gross_earnings = float(sum_result)
    except Exception as e:
        print(f"Error menghitung pendapatan harian untuk print: {e}")

    # 5. Hitung Margin Hari Ini
    margin = MarginSetting.query.first()
    if not margin:
        margin_nominal = 5000
        margin_potongan = 500
    else:
        margin_nominal = margin.nominal_per_margin
        margin_potongan = margin.potongan_per_margin

    margin_count = int(gross_earnings // margin_nominal)
    margin_total = margin_count * margin_potongan

    # 6. Hitung Pendapatan Bersih
    total_earnings = gross_earnings - margin_total

    # 7. Ambil Daftar Transaksi Pembayaran Diterima Hari Ini
    daily_transactions = []
    try:
        daily_transactions = Transaction.query.options(
            db.joinedload(Transaction.user)
        ).filter(
            Transaction.related_user_id == current_user.id,
            Transaction.type == 'payment',
            Transaction.timestamp >= today_start_utc_for_calc,
            Transaction.timestamp < tomorrow_start_utc_for_calc
        ).order_by(Transaction.timestamp.asc()).all()
    except Exception as e:
        print(f"Error mengambil transaksi harian untuk print: {e}")
        flash("Gagal mengambil rincian transaksi untuk dicetak.", "warning")

    # 8. Siapkan data untuk dikirim ke template cetak
    report_data = {
        "vendor_name": current_user.username,
        "report_date_str": today_wib.strftime('%A, %d %B %Y'),
        "total_earnings": total_earnings,      # bersih
        "gross_earnings": gross_earnings,      # kotor
        "margin_total": margin_total,          # margin
        "transactions": daily_transactions,
        "print_time_wib": format_wib(datetime.utcnow())
    }

    # 9. Render template khusus untuk cetak
    html = render_template('vendor_daily_print.html', **report_data)
    pdf = HTML(string=html, base_url=request.base_url).write_pdf()
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=laporan_harian.pdf'
    return response


@app.route('/admin/users/toggle_active/<int:user_id>', methods=['POST'])
@login_required
def toggle_user_active(user_id):
    # 1. Otorisasi: Pastikan hanya admin
    if current_user.role != 'admin':
        flash("Anda tidak memiliki izin untuk melakukan aksi ini.", "danger")
        # Arahkan ke dashboard biasa jika bukan admin mencoba akses
        return redirect(url_for('dashboard'))

    # 2. Cari user yang akan diubah statusnya
    # get_or_404 akan otomatis menampilkan halaman Not Found jika ID tidak ada
    user_to_toggle = User.query.get_or_404(user_id)

    # 3. Validasi: Admin tidak bisa menonaktifkan diri sendiri
    if user_to_toggle.id == current_user.id:
        flash("Anda tidak dapat mengubah status aktif akun Anda sendiri.", "warning")
        return redirect(url_for('admin_users'))

    # 4. Ubah status is_active
    try:
        # Balikkan statusnya (True jadi False, False jadi True)
        user_to_toggle.is_active = not user_to_toggle.is_active
        # Tentukan pesan berdasarkan status baru
        status_text = "diaktifkan" if user_to_toggle.is_active else "dinonaktifkan"
        # Simpan perubahan ke database
        db.session.commit()
        flash(
            f"Akun '{user_to_toggle.username}' berhasil {status_text}.", "success")
    except Exception as e:
        # Jika gagal simpan, batalkan perubahan
        db.session.rollback()
        flash(f"Gagal mengubah status akun: {str(e)}", "danger")

    # 5. Redirect kembali ke halaman daftar pengguna
    return redirect(url_for('admin_users'))


@app.route('/admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_user(user_id):
    # Otorisasi Admin
    if current_user.role != 'admin':
        flash("Anda tidak memiliki izin.", "danger")
        return redirect(url_for('dashboard'))

    # Ambil data user yang mau diedit
    user_to_edit = User.query.get_or_404(user_id)

    if request.method == 'POST':
        # Ambil data baru dari form
        new_id = request.form.get('id')
        new_username = request.form.get('username').strip()
        new_role = request.form.get('role')
        new_kelas = request.form.get('kelas')
        new_jurusan = request.form.get('jurusan')
        new_password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        error = False

        # Validasi ID
        if not new_id or not new_id.isdigit() or len(new_id) != 5:
            flash("ID harus terdiri dari 5 angka.", "danger")
            error = True
        else:
            # Cek keunikan ID jika diubah
            if str(user_to_edit.id) != new_id:
                existing_id = User.query.filter(User.id == int(new_id), User.id != user_id).first()
                if existing_id:
                    flash("ID sudah digunakan oleh pengguna lain.", "danger")
                    error = True

        # Validasi username
        if not new_username:
            flash("Nama lengkap tidak boleh kosong.", "danger")
            error = True

        # Validasi role
        if user_to_edit.role != 'admin':
            if not new_role or new_role not in ['siswa', 'penjual']:
                flash("Peran yang dipilih tidak valid.", "danger")
                error = True
        else:
            new_role = 'admin'

        # Validasi kelas/jurusan jika siswa
        kelas_final = None
        jurusan_final = None
        if new_role == 'siswa':
            kelas_valid = ["X", "XI", "XII"]
            jurusan_valid = ["RPL", "DKV1", "DKV2", "AK", "MP", "BR"]
            if not new_kelas or new_kelas not in kelas_valid:
                flash("Pilihan Kelas untuk siswa tidak valid.", "danger")
                error = True
            if not new_jurusan or new_jurusan not in jurusan_valid:
                flash("Pilihan Jurusan untuk siswa tidak valid.", "danger")
                error = True
            kelas_final = new_kelas
            jurusan_final = new_jurusan

        # Cek keunikan username jika diubah
        if new_username != user_to_edit.username:
            existing_user = User.query.filter(
                User.username == new_username, User.id != user_id).first()
            if existing_user:
                flash(f"Nama '{new_username}' sudah digunakan.", "danger")
                error = True

        # Validasi password jika ingin diubah
        if new_password:
            if len(new_password) < 8:
                flash("Password minimal 8 karakter.", "danger")
                error = True
            if new_password != confirm_password:
                flash("Password dan konfirmasi password tidak cocok.", "danger")
                error = True

        if not error:
            try:
                # Update ID jika berubah
                if str(user_to_edit.id) != new_id:
                    user_to_edit.id = int(new_id)
                user_to_edit.username = new_username
                user_to_edit.role = new_role
                user_to_edit.kelas = kelas_final
                user_to_edit.jurusan = jurusan_final
                if new_password:
                    user_to_edit.set_password(new_password)
                db.session.commit()
                flash(f"Data pengguna '{new_username}' berhasil diperbarui.", "success")
                return redirect(url_for('admin_users'))
            except Exception as e:
                db.session.rollback()
                flash(f"Gagal memperbarui data: {str(e)}", "danger")

    return render_template('admin/admin_edit_users.html', user_to_edit=user_to_edit)


@app.route('/history')
@login_required
def history():
    # Pastikan hanya siswa
    if current_user.role != 'siswa':
        return redirect(url_for('dashboard'))

    # TODO: Salin dan sesuaikan query + filter transaksi dari fungsi dashboard lama
    # Query Sederhana Awal:
    transactions = Transaction.query.filter(
        or_(Transaction.user_id == current_user.id,
            and_(Transaction.related_user_id == current_user.id, Transaction.type == 'topup'))
    ).order_by(Transaction.timestamp.desc()).all()

    return render_template('history.html', transactions=transactions)


@app.route('/pay_vendor', methods=['GET', 'POST'])
@login_required
def pay_vendor():
    if current_user.role != 'siswa':
        logger.warning(
            f"Unauthorized access to /pay_vendor by user {current_user.username}, role: {current_user.role}")
        flash("Hanya siswa yang dapat mengakses halaman pembayaran.", "danger")
        return redirect(url_for('dashboard'))

    # CEK APP LOCK DI AWAL (block baik GET maupun POST)
    ok, msg = ensure_app_unlocked()
    if not ok:
        flash(msg, 'warning')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        vendor_id = request.form.get('vendor_id')
        amount_str = request.form.get('amount')
        logger.debug(
            f"Processing payment: vendor_id={vendor_id}, amount={amount_str}")

        vendor = User.query.get(vendor_id)
        amount = 0.0
        error_message = None

        # --- Validasi Input ---
        if not vendor or vendor.role != 'penjual':
            error_message = "Penjual tidak valid atau tidak ditemukan."
        elif not amount_str:
            error_message = "Jumlah pembayaran wajib diisi."
        else:
            try:
                amount = float(amount_str)
                if amount <= 0:
                    error_message = "Jumlah pembayaran harus positif."
            except ValueError:
                error_message = "Jumlah pembayaran harus berupa angka."

        # --- Validasi Saldo ---
        if not error_message:
            db.session.refresh(current_user)
            if current_user.balance < amount:
                error_message = f"Saldo Anda tidak mencukupi! Saldo saat ini: {format_rupiah(current_user.balance)}"

        # --- Handle jika ada error validasi ---
        if error_message:
            logger.warning(f"Payment validation failed: {error_message}")
            flash(error_message, "danger")
            vendors = User.query.filter_by(
                role='penjual', is_active=True).order_by(User.username).all()
            return render_template('/siswa/pay_vendor.html', vendors=vendors, active_page='pay_vendor')

        # --- Proses Transaksi & Notifikasi ---
        try:
            # Ambil saldo sebelum
            student_balance_before = current_user.balance
            vendor_balance_before = vendor.balance

            # Update saldo
            current_user.balance -= amount
            vendor.balance += amount
            student_balance_after = current_user.balance

            # 1. Add Transaksi Pembayaran ke session
            payment_transaction = Transaction(
                user_id=current_user.id,
                related_user_id=vendor.id,
                type='payment',
                amount=amount,
                user_balance_before=student_balance_before,
                user_balance_after=student_balance_after,
                description=f"Pembayaran ke penjual: {vendor.username}"
            )
            db.session.add(payment_transaction)

            # 2. Add Notifikasi Pembayaran ke session
            notif_message = f"Pembayaran sebesar {format_rupiah(amount)} ke {vendor.username} berhasil."
            new_notif = Notification(
                user_id=current_user.id, message=notif_message)
            db.session.add(new_notif)

            # 3. Check and unlock new borders (termasuk border jurusan)
            check_and_unlock_borders(current_user.id)

            # 4. Commit SEMUA perubahan sekaligus
            db.session.commit()
            logger.info(
                f"Payment successful: user={current_user.username}, vendor={vendor.username}, amount={amount}")

            flash(
                f"Pembayaran sebesar {format_rupiah(amount)} kepada {vendor.username} BERHASIL!", "success")
            return redirect(url_for('dashboard'))

        except Exception as e:
            db.session.rollback()
            logger.error(
                f"Error processing payment for user {current_user.username}: {e}")
            flash("GAGAL melakukan pembayaran: Terjadi kesalahan sistem.", "danger")
            return redirect(url_for('pay_vendor'))

    # --- Method GET ---
    vendors = User.query.filter_by(
        role='penjual', is_active=True).order_by(User.username).all()
    logger.debug(
        f"Rendering /siswa/pay_vendor.html with {len(vendors)} vendors")
    return render_template('/siswa/pay_vendor.html', vendors=vendors, active_page='pay_vendor')


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        if 'profile_pic' not in request.files:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'No file uploaded'})
            flash('Tidak ada file yang dipilih', 'danger')
            return redirect(url_for('profile'))

        file = request.files['profile_pic']
        if file.filename == '':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'No file selected'})
            flash('Tidak ada file yang dipilih', 'danger')
            return redirect(url_for('profile'))

        if file and allowed_file(file.filename):
            try:
                # Hapus foto lama jika ada
                if current_user.profile_pic_filename:
                    old_pic = os.path.join(app.config['UPLOAD_FOLDER'], current_user.profile_pic_filename)
                    if os.path.exists(old_pic):
                        os.remove(old_pic)

                # Generate unique filename
                timestamp = int(datetime.now().timestamp())
                filename = secure_filename(f"{current_user.id}_{timestamp}.jpg")
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

                # Simpan file
                file.save(file_path)

                # Update database
                current_user.profile_pic_filename = filename
                db.session.commit()

                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({
                        'success': True,
                        'message': 'Profile picture updated successfully'
                    })

                flash('Foto profil berhasil diperbarui', 'success')
                return redirect(url_for('profile'))

            except Exception as e:
                db.session.rollback()
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({
                        'success': False,
                        'message': f'Error: {str(e)}'
                    })
                flash('Gagal mengupload foto profil', 'danger')
                return redirect(url_for('profile'))

        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': False,
                    'message': 'Invalid file type'
                })
            flash('Tipe file tidak diizinkan', 'danger')
            return redirect(url_for('profile'))

    # Ambil semua border yang tersedia
    available_borders = Border.query.all()

    # Hitung total transaksi user
    total_transactions = Transaction.query.filter(
        or_(
            Transaction.user_id == current_user.id,
            and_(
                Transaction.related_user_id == current_user.id,
                Transaction.type == 'payment'
            )
        )
    ).count()

    # Cek border yang sudah terbuka berdasarkan total transaksi dan jurusan
    unlocked_border_ids = []
    for border in available_borders:
        # Border tanpa jurusan (umum) atau sesuai jurusan user
        if (border.jurusan is None or border.jurusan == current_user.jurusan) and \
           total_transactions >= border.required_transactions:
            user_border = UserBorder.query.filter_by(
                user_id=current_user.id,
                border_id=border.id
            ).first()
            
            if not user_border:
                user_border = UserBorder(
                    user_id=current_user.id,
                    border_id=border.id,
                    unlocked_at=datetime.utcnow()
                )
                db.session.add(user_border)
                
            unlocked_border_ids.append(border.id)
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Terjadi kesalahan saat memperbarui status border: {str(e)}", "danger")


    filename = current_user.profile_pic_filename if current_user.profile_pic_filename else 'default.png'

    return render_template('siswa/profile.html',
                         available_borders=available_borders,
                         unlocked_border_ids=unlocked_border_ids,
                         total_transactions=total_transactions,
                         active_page='profile',
                         profile_pic_filename=filename  # <-- Tambahkan ini!
                         )

# Tambahkan route baru

@app.route('/pengaturan_profile')
@login_required
def pengaturan_profile():
    return render_template('siswa/pengaturan_profile.html', active_page='pengaturan_profile')


@app.route('/change_border/<int:border_id>', methods=['POST'])
@login_required
def change_border(border_id):
    try:
        border = Border.query.get_or_404(border_id)
        
        # Jika request untuk melepas border
        if 'remove_border' in request.form:
            # Gunakan text() untuk raw SQL
            db.session.execute(
                db.text("UPDATE user SET active_border_id = NULL WHERE id = :user_id"),
                {"user_id": current_user.id}
            )
            current_user.active_border_id = None
            current_user.active_border = None
            flash("Border berhasil dilepas!", "success")
        else:
            # Set border baru
            current_user.active_border_id = border_id
            current_user.active_border = border
            flash(f"Border berhasil diubah ke {border.name}!", "success")
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        flash(f"Gagal mengubah border: {str(e)}", "danger")
    
    return redirect(url_for('profile'))


@app.route('/riwayat_siswa')
@login_required
def riwayat_siswa():
    if current_user.role != 'siswa':
        flash("Halaman ini hanya untuk siswa.", "danger")
        return redirect(url_for('dashboard'))

    # Ambil filter bulan dan tahun dari request
    filter_month = request.args.get('filter_month', None, type=int)
    filter_year = request.args.get('filter_year', None, type=int)
    selected_type_filter = request.args.get('filter_type', 'all')

    # Zona Waktu WIB
    try:
        wib = ZoneInfo("Asia/Jakarta")
    except Exception:
        wib = None
    now_wib = datetime.now(wib) if wib and isinstance(wib, ZoneInfo) else datetime.utcnow() + timedelta(hours=7)
    today_wib = now_wib.date()

    # Default: bulan dan tahun saat ini jika tidak dipilih
    if not filter_month:
        filter_month = today_wib.month
    if not filter_year:
        filter_year = today_wib.year

    # Hitung tanggal awal dan akhir bulan
    start_dt = date(filter_year, filter_month, 1)
    # Hitung hari terakhir bulan
    if filter_month == 12:
        end_dt = date(filter_year + 1, 1, 1) - timedelta(days=1)
    else:
        end_dt = date(filter_year, filter_month + 1, 1) - timedelta(days=1)

    # Konversi ke UTC
    start_wib_day_local = datetime.combine(start_dt, time.min)
    if wib and isinstance(wib, ZoneInfo):
        start_wib_day_local = start_wib_day_local.replace(tzinfo=wib)
    end_wib_day_local = datetime.combine(end_dt, time.max)
    if wib and isinstance(wib, ZoneInfo):
        end_wib_day_local = end_wib_day_local.replace(tzinfo=wib)
    start_utc_for_query = start_wib_day_local.astimezone(ZoneInfo("UTC")) if wib and isinstance(wib, ZoneInfo) else start_wib_day_local - timedelta(hours=7)
    end_utc_for_query = end_wib_day_local.astimezone(ZoneInfo("UTC")) if wib and isinstance(wib, ZoneInfo) else end_wib_day_local - timedelta(hours=7)

    # Query transaksi sesuai bulan & tahun
    base_query = Transaction.query.options(
        db.joinedload(Transaction.user), db.joinedload(Transaction.related_user)
    ).filter(
        or_(
            Transaction.user_id == current_user.id,
            Transaction.related_user_id == current_user.id
        ),
        Transaction.timestamp >= start_utc_for_query,
        Transaction.timestamp <= end_utc_for_query
    )

    # Filter tipe transaksi (tambahkan 'transfer' sebagai pilihan)
    if selected_type_filter != 'all':
        if selected_type_filter == 'transfer':
            base_query = base_query.filter(Transaction.type == 'transfer')
        else:
            base_query = base_query.filter(Transaction.type == selected_type_filter)

    transactions = base_query.order_by(Transaction.timestamp.desc()).all()

    bulan_ini = start_dt.strftime('%B')

    # Hitung total bulanan dengan logika benar
    total_bulan = 0
    for tx in transactions:
        if tx.type == 'topup' and tx.related_user_id == current_user.id:
            total_bulan += tx.amount
        elif tx.type == 'transfer':
            if tx.user_id == current_user.id:
                total_bulan -= tx.amount  # transfer keluar
            elif tx.related_user_id == current_user.id:
                total_bulan += tx.amount  # transfer masuk
        elif tx.type == 'payment' and tx.user_id == current_user.id:
            total_bulan -= tx.amount  # pembayaran keluar

    # Render template
    return render_template('/siswa/transaksi_siswa.html',
                           transactions=transactions,
                           filter_month=filter_month,
                           filter_year=filter_year,
                           selected_type_filter=selected_type_filter,
                           bulan_ini=bulan_ini,
                           total_bulan=total_bulan,
                           active_page='riwayat_siswa')
# === Akhir Route Profil ===


@app.route('/request_topup', methods=['GET', 'POST'])
@login_required
def request_topup():
    # Pastikan hanya siswa
    if current_user.role != 'siswa':
        flash("Hanya siswa yang dapat membuat request top up.", "danger")
        return redirect(url_for('dashboard'))

    # Cek apakah siswa ini sudah punya request yang statusnya 'pending'
    pending_request = TopUpRequest.query.filter_by(
        student_id=current_user.id, status='pending').first()

    # Proses form jika method POST
    if request.method == 'POST':
        print(f"Received POST request to /request_topup from user {current_user.id}: {request.form}")

        # Jika sudah ada pending, jangan izinkan buat request baru
        if pending_request:
            flash('Anda sudah memiliki 1 request top up yang belum diproses. Harap selesaikan pembayaran tunai di Bank Mini dahulu.', 'warning')
            return redirect(url_for('request_topup'))

        # Ambil jumlah dari form
        amount_str = request.form.get('amount')
        amount = 0.0
        error = False

        # Ambil file bukti transfer dari form
        bukti_file = request.files.get('bukti_transfer')
        bukti_filename = None

        if not amount_str:
            flash('Jumlah request top up wajib diisi.', 'danger')
            error = True
        else:
            try:
                amount = float(amount_str)
                if amount < 1000:
                    flash('Jumlah minimum request top up adalah Rp 1.000.', 'danger')
                    error = True
            except ValueError:
                flash('Jumlah harus berupa angka.', 'danger')
                error = True

        # Validasi file bukti transfer jika di-upload
        if bukti_file and bukti_file.filename != '':
            if allowed_file(bukti_file.filename):
                bukti_filename = secure_filename(f"bukti_{current_user.id}_{int(datetime.now().timestamp())}.jpg")
                bukti_path = os.path.join(app.config['UPLOAD_FOLDER'], bukti_filename)
                bukti_file.save(bukti_path)
            else:
                flash('Tipe file bukti transfer tidak diizinkan. Hanya jpg/png/gif.', 'danger')
                error = True

        # Jika validasi lolos
        if not error:
            try:
                # Buat record request baru
                new_request = TopUpRequest(
                    student_id=current_user.id,
                    amount=amount,
                    status='pending',
                    request_timestamp=datetime.now(pytz.timezone("UTC")),
                    bukti_transfer_filename=bukti_filename
                )
                db.session.add(new_request)
                db.session.commit()
                flash(
                    f'Request top up sebesar {format_rupiah(amount)} berhasil dibuat. Silakan lakukan pembayaran tunai sejumlah tersebut di Bank Mini.', 'success')
                return redirect(url_for('request_topup'))
            except Exception as e:
                db.session.rollback()
                flash('Gagal membuat request top up, silakan coba lagi.', 'danger')
                return redirect(url_for('request_topup'))

        # Jika validasi gagal, redirect kembali ke GET request
        return redirect(url_for('request_topup'))

    # Method GET: Tampilkan form dan riwayat request
    else:
        recent_requests = TopUpRequest.query.filter_by(student_id=current_user.id)\
            .order_by(TopUpRequest.request_timestamp.desc())\
            .limit(5).all()

        return render_template('/siswa/topup_siswa.html',
                               pending_request=pending_request,
                               recent_requests=recent_requests)
# ==========================================


@app.route('/admin/pending_topups')
@login_required
def admin_pending_topups():
    # Otorisasi: Pastikan hanya admin
    if current_user.role != 'admin':
        flash("Akses ditolak. Fitur ini hanya untuk admin.", "danger")
        return redirect(url_for('dashboard'))

    try:
        # Ambil semua data TopUpRequest yang statusnya 'pending'
        # Sekaligus ambil data siswa terkait (joinedload untuk efisiensi)
        # Urutkan berdasarkan request paling lama (ascending)
        pending_requests = TopUpRequest.query.options(
            db.joinedload(TopUpRequest.student)  # Eager load data siswa
        ).filter_by(status='pending').order_by(TopUpRequest.request_timestamp.asc()).all()
    except Exception as e:
        print(f"Error mengambil request topup pending: {e}")
        flash("Gagal memuat daftar request top up.", "danger")
        pending_requests = []  # List kosong jika error

    # Render template baru, kirim data request pending
    return render_template('/admin/admin_pending_topups.html', requests=pending_requests, active_page='admin_pending_topups')


@app.route('/admin/approve_topup/<int:request_id>', methods=['POST'])
@login_required
def admin_approve_topup(request_id):
    # 1. Otorisasi: Pastikan hanya admin
    if current_user.role != 'admin':
        logger.warning(
            f"Unauthorized access to /admin/approve_topup/{request_id} by user {current_user.username}")
        flash("Anda tidak memiliki izin.", "danger")
        return redirect(url_for('dashboard'))

    # 2. Cari request top up berdasarkan ID
    topup_request = TopUpRequest.query.get_or_404(request_id)
    logger.debug(
        f"Processing top-up request ID {request_id}, status: {topup_request.status}")

    # 3. Validasi Status: Pastikan masih 'pending'
    if topup_request.status != 'pending':
        logger.info(
            f"Top-up request ID {request_id} already processed, status: {topup_request.status}")
        flash(
            f"Request top up ini sudah diproses sebelumnya (Status: {topup_request.status}).", "warning")
        return redirect(url_for('admin_pending_topups'))

    # 4. Cari data siswa
    student = User.query.get(topup_request.student_id)
    if not student:
        logger.error(f"Student not found for top-up request ID {request_id}")
        flash(
            f"Data siswa untuk request ID {request_id} tidak ditemukan.", "danger")
        return redirect(url_for('admin_pending_topups'))

    # 5. Proses Inti: Update Saldo, Buat Transaksi, Update Request, Buat Notifikasi
    try:
        amount_to_add = topup_request.amount
        balance_before = student.balance
        student.balance += amount_to_add  # Tambah saldo siswa
        balance_after = student.balance

        # Buat record transaksi 'topup'
        new_transaction = Transaction(
            user_id=current_user.id,        # ID Admin
            related_user_id=student.id,     # ID Siswa
            type='topup',
            amount=amount_to_add,
            user_balance_before=balance_before,
            user_balance_after=balance_after,
            description=f"Persetujuan Request Top Up ID: {topup_request.id}"
        )
        db.session.add(new_transaction)

        # Update status record TopUpRequest
        topup_request.status = 'approved'
        topup_request.admin_id = current_user.id
        topup_request.processed_timestamp = datetime.utcnow()
        db.session.add(topup_request)

        # Buat Notifikasi untuk Siswa
        notif_message = f"Top up sebesar {format_rupiah(amount_to_add)} telah disetujui oleh admin."
        new_notif = Notification(user_id=student.id, message=notif_message)
        db.session.add(new_notif)

        # Commit semua perubahan
        db.session.commit()
        logger.info(
            f"Top-up request ID {request_id} approved for student {student.username}, amount: {amount_to_add}")

        # Beri pesan sukses
        flash(
            f"Request Top Up untuk {student.username} sebesar {format_rupiah(amount_to_add)} berhasil disetujui.", "success")

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error approving top-up request ID {request_id}: {e}")
        flash("GAGAL menyetujui request top up: Terjadi kesalahan sistem.", "danger")

    return redirect(url_for('admin_pending_topups'))

# === Akhir Fungsi Route Approve Top Up ===


@app.route('/notifications/mark_read', methods=['POST'])
@login_required
def mark_notifications_read():
    # Hanya user yg login yg bisa tandai notifnya sendiri
    if not current_user.is_authenticated:
        return jsonify(success=False, message="User not logged in"), 401

    try:
        # Cari semua notifikasi milik user ini yang belum dibaca
        unread_notifications = Notification.query.filter_by(
            user_id=current_user.id, is_read=False).all()

        # Jika ada yang belum dibaca, ubah statusnya
        if unread_notifications:
            for notif in unread_notifications:
                notif.is_read = True
            # Commit perubahan ke database
            db.session.commit()
            print(
                f"Marked {len(unread_notifications)} notifications as read for user {current_user.id}")
        else:
            print(
                f"No unread notifications to mark as read for user {current_user.id}")

        # Kirim response sukses ke Javascript
        return jsonify(success=True)

    except Exception as e:
        db.session.rollback()
        print(
            f"Error marking notifications as read for user {current_user.id}: {e}")
        # Kirim response gagal ke Javascript
        return jsonify(success=False, message="Gagal update status notifikasi"), 500
# =================================================


@app.route('/detail_notif')
@login_required
def detail_notif():
    # Pastikan hanya siswa
    if current_user.role != 'siswa':
        flash("Halaman ini hanya untuk siswa.", "danger")
        return redirect(url_for('dashboard'))

    # Ambil SEMUA notifikasi user ini (sudah diurut terbaru di atas oleh backref)
    all_notifications = current_user.notifications.all()

    return render_template('/siswa/detail_notif.html', notifications=all_notifications, active_page='detail_notif')
    # === BARU: Model untuk Notifikasi Pengguna ===


@app.route('/admin/register_user', methods=['GET', 'POST'])
@login_required
def admin_register_user():
    # Hanya admin yang boleh akses
    if current_user.role != 'admin':
        flash("Hanya admin yang dapat mendaftarkan akun baru.", "danger")
        return redirect(url_for('dashboard'))

    # --- Handle Import Excel ---
    if request.method == 'POST' and 'file' in request.files:
        role = request.form.get('role')
        file = request.files.get('file')
        if not file or not role:
            flash("File dan jenis user wajib diisi.", "danger")
            return redirect(url_for('admin_register_user'))

        import openpyxl
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        imported = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if role == 'siswa':
                id, username, kelas, jurusan = row
                password = secrets.token_urlsafe(8)[:8]
                if User.query.get(id):
                    continue
                user = User(id=id, username=username, role='siswa', kelas=kelas, jurusan=jurusan)
            else:
                id, username = row
                password = secrets.token_urlsafe(8)[:8]
                if User.query.get(id):
                    continue
                user = User(id=id, username=username, role='penjual')
            user.set_password(password)
            user.plain_password = password  # Simpan password asli untuk notifikasi
            db.session.add(user)
            imported += 1
        db.session.commit()
        flash(f"Berhasil import {imported} {role}. Password otomatis digenerate.", "success")
        return redirect(url_for('admin_register_user'))

    # --- Handle Registrasi Manual ---
    if request.method == 'POST' and 'file' not in request.files:
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        role = request.form.get('role')
        kelas = request.form.get('kelas')
        jurusan = request.form.get('jurusan')

        error = False
        # Validasi dasar
        if not username or not password or not confirm_password or not role:
            flash('Semua kolom wajib diisi!', 'danger')
            error = True
        if password != confirm_password:
            flash('Password dan konfirmasi password tidak cocok!', 'danger')
            error = True
        if role not in ['siswa', 'penjual']:
            flash('Peran yang dipilih tidak valid.', 'danger')
            error = True

        # Validasi kelas/jurusan jika siswa
        if role == 'siswa':
            kelas_valid = ["X", "XI", "XII"]
            jurusan_valid = ["RPL", "DKV1", "DKV2", "AK", "MP", "BR"]
            if not kelas or not jurusan:
                flash('Siswa wajib memilih Kelas dan Jurusan!', 'danger')
                error = True
            if kelas not in kelas_valid:
                flash('Kelas yang dipilih tidak valid.', 'danger')
                error = True
            if jurusan not in jurusan_valid:
                flash('Jurusan yang dipilih tidak valid.', 'danger')
                error = True

        # Cek username unik
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Username ini sudah digunakan, silakan pilih username lain.', 'warning')
            error = True

        if error:
            return redirect(url_for('admin_register_user'))

        # Buat user baru
        user_data = {
            'username': username,
            'role': role
        }
        if role == 'siswa':
            user_data['kelas'] = kelas
            user_data['jurusan'] = jurusan

        new_user = User(**user_data)
        new_user.set_password(password)
        new_user.plain_password = password

        try:
            db.session.add(new_user)
            db.session.commit()
            flash(f'Akun {username} ({role}) berhasil dibuat!', 'success')
            return redirect(url_for('admin_users'))
        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan saat membuat akun: {str(e)}', 'danger')
            return redirect(url_for('admin_register_user'))

    # Method GET: tampilkan form
    return render_template('/admin/admin_register_user.html', active_page='admin_register_user')


@app.route('/top_board')
def top_board():
    # Dapatkan waktu WIB saat ini
    now_wib = datetime.now(wib_tz)
    today_wib = now_wib.date()

    # Tentukan batas waktu UTC untuk hari ini (WIB)
    today_wib_start_local = datetime.combine(
        today_wib, time.min).replace(tzinfo=wib_tz)
    tomorrow_wib_start_local = today_wib_start_local + timedelta(days=1)
    today_start_utc = today_wib_start_local.astimezone(utc_tz)
    tomorrow_start_utc = tomorrow_wib_start_local.astimezone(utc_tz)

    try:
        # Query untuk menghitung total pembayaran per siswa hari ini
        top_spenders = db.session.query(
            User.id,
            User.username,
            User.kelas,
            User.jurusan,
            func.sum(Transaction.amount).label('total_spent')
        ).join(
            Transaction, Transaction.user_id == User.id
        ).filter(
            User.role == 'siswa',
            User.is_active == True,
            Transaction.type == 'payment',
            Transaction.timestamp >= today_start_utc,
            Transaction.timestamp < tomorrow_start_utc
        ).group_by(
            User.id, User.username, User.kelas, User.jurusan
        ).order_by(
            func.sum(Transaction.amount).desc()
        ).limit(10).all()  # Ambil top 10 siswa

        # Format data untuk template
        leaderboard = []
        for rank, spender in enumerate(top_spenders, 1):
            user_obj = User.query.get(spender.id)
            leaderboard.append({
                'rank': rank,
                'username': spender.username,
                'kelas': spender.kelas or '-',
                'jurusan': spender.jurusan or '-',
                'total_spent': format_rupiah(spender.total_spent),
                'profile_pic_filename': user_obj.profile_pic_filename if user_obj and user_obj.profile_pic_filename else 'default.png',
                'active_border': user_obj.active_border if user_obj else None
            })

    except Exception as e:
        print(f"Error fetching top board data: {e}")
        flash("Gagal memuat data Top Board.", "warning")
        leaderboard = []

    # Render template Top Board
    return render_template(
        'top_board.html',
        leaderboard=leaderboard,
        date_str=today_wib.strftime('%A, %d %B %Y'),
        current_user_role=current_user.role if current_user.is_authenticated else None,
        active_page='top_board'
    )


@app.route('/update_password', methods=['POST'])
@login_required
def update_password():
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    # Cek password saat ini menggunakan password_hash
    if not check_password_hash(current_user.password_hash, current_password):
        flash('Password saat ini salah!', 'danger')
        return redirect(url_for('profile'))

    # Cek konfirmasi password
    if new_password != confirm_password:
        flash('Password baru dan konfirmasi tidak cocok!', 'danger')
        return redirect(url_for('profile'))

    # Update password menggunakan method set_password
    try:
        current_user.set_password(new_password)  # Gunakan method set_password
        db.session.commit()
        flash('Password berhasil diupdate!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal update password: {str(e)}', 'danger')

    return redirect(url_for('profile'))

@app.route('/transaksi/<int:transaksi_id>')
@login_required
def detail_transaksi(transaksi_id):
    transaksi = Transaction.query.get_or_404(transaksi_id)
    # Jika transaksi topup, tampilkan related_user_id (siswa)
    if transaksi.type == 'topup' and transaksi.related_user_id:
        user_id_masked = str(transaksi.related_user_id)[:2] + '***'
    else:
        user_id_masked = str(transaksi.user_id)[:2] + '***'
    return render_template('siswa/detail_transaksi.html', transaksi=transaksi, user_id_masked=user_id_masked)


@app.route('/admin/profile')
@login_required
def admin_profile():
    if current_user.role != 'admin':
        flash('Anda tidak memiliki akses ke halaman ini!', 'error')
        return redirect(url_for('dashboard'))
    return render_template('admin/admin_profile.html', active_page='admin_profile')


@app.route('/admin/update-password', methods=['POST'])
@login_required
def admin_update_password():
    if current_user.role != 'admin':
        flash('Anda tidak memiliki akses ke halaman ini!', 'error')
        return redirect(url_for('dashboard'))

    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    # Validasi password saat ini
    if not check_password_hash(current_user.password_hash, current_password):
        flash('Password saat ini salah!', 'error')
        return redirect(url_for('admin_profile'))

    # Validasi password baru
    if new_password != confirm_password:
        flash('Password baru dan konfirmasi tidak cocok!', 'error')
        return redirect(url_for('admin_profile'))

    try:
        current_user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        flash('Password berhasil diupdate!', 'success')
    except:
        db.session.rollback()
        flash('Gagal update password. Silakan coba lagi.', 'error')

    return redirect(url_for('admin_profile'))

@app.route('/admin/laporan')
@login_required
def admin_laporan():
    if current_user.role != 'admin':
        flash("Akses ditolak. Halaman ini hanya untuk admin.", "danger")
        return redirect(url_for('dashboard'))

    filter_status = request.args.get('filter', 'all')
    if filter_status == 'sudah':
        daftar_laporan = Laporan.query.filter(
            (Laporan.ditanggapi == True) | (Laporan.ditanggapi == 1)
        ).order_by(Laporan.timestamp.desc()).all()
    elif filter_status == 'belum':
        daftar_laporan = Laporan.query.filter(
            (Laporan.ditanggapi == False) | (Laporan.ditanggapi == 0)
        ).order_by(Laporan.timestamp.desc()).all()
    else:
        daftar_laporan = Laporan.query.order_by(Laporan.timestamp.desc()).all()

    return render_template(
        'admin/admin_laporan.html',
        daftar_laporan=daftar_laporan,
        active_page='admin_laporan'
    )

@app.route('/transfer', methods=['GET', 'POST'])
@login_required
def transfer_user():

    # CEK APP LOCK DI AWAL (block baik GET maupun POST)
    ok, msg = ensure_app_unlocked()
    if not ok:
        flash(msg, 'warning')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        target_id = request.form.get('target_id')
        amount = float(request.form.get('amount', 0))
        # Validasi
        if not target_id or amount <= 0:
            flash('ID tujuan dan jumlah harus diisi!', 'danger')
            return redirect(url_for('transfer_user'))
        if amount > current_user.balance:
            flash('Saldo tidak cukup!', 'danger')
            return redirect(url_for('transfer_user'))
        target_user = User.query.filter_by(id=target_id).first()
        if not target_user or target_user.id == current_user.id:
            flash('User tujuan tidak valid!', 'danger')
            return redirect(url_for('transfer_user'))
        # Proses transfer
        user_balance_before = current_user.balance
        current_user.balance -= amount
        target_user.balance += amount
        user_balance_after = current_user.balance
        # Simpan transaksi
        tx = Transaction(
            user_id=current_user.id,
            related_user_id=target_user.id,
            type='transfer',
            amount=amount,
            user_balance_before=user_balance_before,
            user_balance_after=user_balance_after,
            description=f"Transfer ke {target_user.username}"
        )
        db.session.add(tx)
        # Buat notifikasi untuk pengirim
        notif_sender = Notification(
            user_id=current_user.id,
            message=f"Transfer ke {target_user.username} sebesar {amount:,.0f} berhasil.",
            timestamp=datetime.utcnow(),
            is_read=False
        )
        db.session.add(notif_sender)
        # Buat notifikasi untuk penerima
        notif_receiver = Notification(
            user_id=target_user.id,
            message=f"Kamu menerima transfer dari {current_user.username} sebesar {amount:,.0f}.",
            timestamp=datetime.utcnow(),
            is_read=False
        )
        db.session.add(notif_receiver)
        db.session.commit()
        flash('Transfer berhasil!', 'success')
        return redirect(url_for('dashboard'))
    return render_template('siswa/transfer_user.html')

@app.route('/api/username/<user_id>')
@login_required
def api_username(user_id):
    user = User.query.filter_by(id=user_id).first()
    if user:
        photo_url = None
        border_url = None
        # Foto profil
        if user.profile_pic_filename:
            photo_url = url_for('static', filename=f'profile_pics/{user.profile_pic_filename}')
        # Border aktif
        if user.active_border and user.active_border.image_path:
            border_url = url_for('static', filename=user.active_border.image_path)
        return {
            'username': user.username,
            'photo_url': photo_url,
            'border_url': border_url
        }
    return {'username': None}

@app.route('/admin/laporan/tanggapi/<int:laporan_id>', methods=['POST'])
@login_required
def tanggapi_laporan(laporan_id):
    if current_user.role != 'admin':
        flash("Akses ditolak.", "danger")
        return redirect(url_for('admin_laporan'))
    laporan = Laporan.query.get_or_404(laporan_id)
    laporan.ditanggapi = True
    db.session.commit()
    flash("Laporan telah ditandai sebagai sudah ditanggapi.", "success")
    return redirect(url_for('admin_laporan'))

@app.route('/buat_laporan', methods=['GET', 'POST'])
@login_required
def buat_laporan():
    if request.method == 'POST':
        keterangan = request.form.get('keterangan')
        file = request.files.get('bukti_laporan')
        filename = None

        if file and file.filename != '':
            ext = file.filename.rsplit('.', 1)[-1].lower()
            if ext in ['jpg', 'jpeg', 'png', 'gif', 'mp4', 'mov', 'avi']:
                filename = secure_filename(f"laporan_{current_user.id}_{int(datetime.now().timestamp())}.{ext}")
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            else:
                flash('Tipe file tidak didukung. Hanya gambar/video.', 'danger')
                return redirect(url_for('buat_laporan'))

        # Simpan laporan ke database
        laporan = Laporan(
            user_id=current_user.id,
            bukti_filename=filename,
            keterangan=keterangan,
            timestamp=datetime.utcnow()
        )
        db.session.add(laporan)
        db.session.commit()

        flash('Laporan berhasil dikirim!', 'success')
        return redirect(url_for('buat_laporan'))  # Redirect ke halaman yang sama agar pesan muncul

    return render_template('siswa/buat_laporan.html', pengirim=current_user.username)

@app.route('/atur_pin', methods=['GET', 'POST'])
@login_required
def atur_pin():
    if request.method == 'POST':
        pin = request.form.get('pin')
        old_pin = request.form.get('old_pin')
        # Validasi: harus 6 digit angka
        if not pin or not pin.isdigit() or len(pin) != 6:
            flash('PIN harus 6 digit angka!', 'danger')
            return redirect(url_for('atur_pin'))
        # Jika sudah punya PIN, cek PIN lama dulu
        if current_user.pin_transfer:
            if not old_pin or old_pin != current_user.pin_transfer:
                flash('PIN lama salah!', 'danger')
                return redirect(url_for('atur_pin'))
        current_user.pin_transfer = pin
        db.session.commit()
        flash('PIN transfer berhasil disimpan!', 'success')
        return redirect(url_for('profile'))
    return render_template('siswa/atur_pin.html')

@app.route('/api/cek_pin', methods=['POST'])
@login_required
def api_cek_pin():
    data = request.get_json()
    pin = data.get('pin')
    valid = current_user.pin_transfer == pin
    return jsonify({'valid': valid})



@app.route('/admin/download_template/<role>')
@login_required
def download_template(role):
    wb = openpyxl.Workbook()
    ws = wb.active
    if role == 'siswa':
        ws.append(['id', 'username', 'kelas', 'jurusan'])  # Header
        ws.append(['12345', 'Budi Santoso', 'X', 'RPL'])   # Contoh data siswa 1
        ws.append(['12346', 'Siti Aminah', 'XI', 'DKV1'])  # Contoh data siswa 2
        ws.append(['12347', 'Andi Wijaya', 'XII', 'AK'])   # Contoh data siswa 3
    else:
        ws.append(['id', 'username'])                      # Header
        ws.append(['20001', 'Warung Pak Dedi'])            # Contoh penjual 1
        ws.append(['20002', 'Kantin Bu Sari'])             # Contoh penjual 2

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"template_{role}.xlsx"
    return send_file(stream, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/export_users')
@login_required
def admin_export_users():
    if current_user.role != 'admin':
        flash("Akses ditolak.", "danger")
        return redirect(url_for('admin_users'))

    role = request.args.get('role')
    kelas = request.args.get('kelas')
    jurusan = request.args.get('jurusan')

    query = User.query
    if role:
        query = query.filter_by(role=role)
    if kelas:
        query = query.filter_by(kelas=kelas)
    if jurusan:
        query = query.filter_by(jurusan=jurusan)

    users = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pengguna"
    ws.append(['ID', 'Username', 'Password', 'Role', 'Kelas', 'Jurusan'])

    for user in users:
        # Ambil password asli jika disimpan, atau tampilkan info jika tidak bisa
        password = getattr(user, 'plain_password', None)
        if not password:
            password = '(tidak tersedia)'
        ws.append([
            user.id,
            user.username,
            password,
            user.role,
            user.kelas if user.kelas else '',
            user.jurusan if user.jurusan else ''
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "export_pengguna.xlsx"
    return send_file(stream, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/verify_device')
@login_required
def verify_device():
    # Jika user belum aktifkan WebAuthn, langsung ke dashboard
    if not current_user.webauthn_enabled:
        return redirect(url_for('dashboard'))
    return render_template('siswa/verify_device.html')

@app.route('/admin/margin', methods=['GET', 'POST'])
@login_required
def admin_margin():
    if current_user.role != 'admin':
        flash("Akses ditolak.", "danger")
        return redirect(url_for('dashboard'))

    margin = MarginSetting.query.first()
    if not margin:
        margin = MarginSetting(nominal_per_margin=5000, potongan_per_margin=500)
        db.session.add(margin)
        db.session.commit()

    if request.method == 'POST':
        nominal = float(request.form.get('nominal_per_margin', 5000))
        potongan = float(request.form.get('potongan_per_margin', 500))
        margin.nominal_per_margin = nominal
        margin.potongan_per_margin = potongan
        db.session.commit()
        flash("Margin berhasil diupdate!", "success")
        return redirect(url_for('admin_margin'))

    return render_template('admin/admin_margin.html', margin=margin)


@app.route('/webauthn/register_challenge', methods=['POST'])
@login_required
def webauthn_register_challenge():
    # Buat challenge dan user info
    challenge = os.urandom(32)
    session['webauthn_challenge'] = base64.b64encode(challenge).decode()
    publicKey = {
        "challenge": base64.b64encode(challenge).decode(),
        "rp": {"name": "Kantin40 Digital"},
        "user": {
            "id": base64.b64encode(str(current_user.id).encode()).decode(),
            "name": current_user.username,
            "displayName": current_user.username
        },
        "pubKeyCredParams": [{"type": "public-key", "alg": -7}],
        "timeout": 60000,
        "attestation": "direct"
    }
    return jsonify({"publicKey": publicKey})

@app.route('/webauthn/register_finish', methods=['POST'])
@login_required
def webauthn_register_finish():
    data = request.get_json()
    current_user.credential_id = data['id']
    current_user.webauthn_enabled = True
    db.session.commit()
    return jsonify({"success": True})

@app.route('/webauthn/login_challenge', methods=['POST'])
@login_required
def webauthn_login_challenge():
    user = current_user
    if not user or not user.credential_id:
        return jsonify({"error": "User belum aktivasi WebAuthn"}), 400
    challenge = os.urandom(32)
    session['webauthn_challenge'] = base64.b64encode(challenge).decode()
    publicKey = {
        "challenge": base64.b64encode(challenge).decode(),
        "allowCredentials": [{
            "type": "public-key",
            "id": base64.b64encode(user.credential_id.encode()).decode()
        }],
        "timeout": 60000
    }
    return jsonify({"publicKey": publicKey})

@app.route('/webauthn/deactivate', methods=['POST'])
@login_required
def webauthn_deactivate():
    current_user.webauthn_enabled = False
    current_user.credential_id = None
    db.session.commit()
    return jsonify({"success": True})

@app.route('/pengaturan_keamanan')
@login_required
def pengaturan_keamanan():
    return render_template('siswa/pengaturan_keamanan.html')

@app.route('/webauthn/login_finish', methods=['POST'])
@login_required
def webauthn_login_finish():
    # Ambil data dari frontend
    data = request.get_json()
    challenge = session.get('webauthn_challenge')
    credential_id = current_user.credential_id

    # --- PSEUDOCODE: Verifikasi credential ---
    # Gunakan library WebAuthn untuk memverifikasi signature, authenticatorData, dsb.
    # Contoh (dengan library webauthn, Anda harus install dan setup dulu):
    # from webauthn import verify_authentication_response
    # try:
    #     verification = verify_authentication_response(
    #         credential_id=credential_id,
    #         client_data_json=data['response']['clientDataJSON'],
    #         authenticator_data=data['response']['authenticatorData'],
    #         signature=data['response']['signature'],
    #         challenge=challenge,
    #         origin='https://kantin40.yourdomain.com',  # Ganti dengan domain Anda
    #         rp_id='kantin40.yourdomain.com',           # Ganti dengan domain Anda
    #         user_handle=data['response'].get('userHandle')
    #     )
    #     if verification.is_successful():
    #         session['device_verified'] = True
    #         return jsonify({"success": True})
    # except Exception as e:
    #     return jsonify({"success": False, "message": str(e)}), 400

    # --- Sementara, anggap selalu sukses ---
    session['device_verified'] = True
    return jsonify({"success": True})



def check_and_unlock_borders(user_id):
    user = User.query.get(user_id)
    if not user:
        return

    # Hitung total transaksi pembayaran user
    total_transactions = Transaction.query.filter(
        Transaction.user_id == user_id,
        Transaction.type == 'payment'
    ).count()

    borders = Border.query.all()
    for border in borders:
        # Border jurusan: hanya cek jika jurusan user sama
        if border.jurusan and border.jurusan != user.jurusan:
            continue

        # Jika jumlah transaksi mencukupi dan border belum dimiliki
        if (total_transactions >= border.required_transactions and
                not UserBorder.query.filter_by(user_id=user_id, border_id=border.id).first()):
            # Buat record UserBorder baru
            new_user_border = UserBorder(
                user_id=user_id,
                border_id=border.id,
                unlocked_at=datetime.utcnow()
            )
            db.session.add(new_user_border)

            # Buat notifikasi untuk user
            notification = Notification(
                user_id=user_id,
                message=f"Selamat! Anda telah membuka border baru: {border.name}!"
            )
            db.session.add(notification)

    try:
        db.session.commit()
    except Exception as e:
        print(f"Error unlocking borders: {e}")
        db.session.rollback()



def add_border_column():
    with app.app_context():
        try:
            # Tambahkan kolom baru jika belum ada
            insp = db.inspect(db.engine)
            has_column = False
            for column in insp.get_columns('border'):
                if column['name'] == 'jurusan':
                    has_column = True
                    break
                    
            if not has_column:
                # Tambah kolom jurusan ke tabel border
                db.session.execute(db.text('ALTER TABLE border ADD COLUMN jurusan VARCHAR(50)'))
                db.session.commit()
                print("Successfully added jurusan column to border table")
                
        except Exception as e:
            db.session.rollback()
            print(f"Error adding jurusan column: {e}")

# ...existing code...

if __name__ == '__main__':
    with app.app_context():
        # DROP dan CREATE khusus tabel AppConfig
        AppConfig.__table__.drop(db.engine, checkfirst=True)
        print("Tabel app_config berhasil dihapus.")
        AppConfig.__table__.create(db.engine, checkfirst=True)
        print("Tabel app_config berhasil dibuat ulang.")

        # Tambah kolom jurusan ke border jika perlu
        add_border_column()
        
    app.run(debug=True, host='0.0.0.0', port=5000)

# ...existing code...


# !!! PENTING: Hapus atau komentari route @app.route('/pay', methods=['POST']) yang lama
# karena logikanya akan dipindah ke /pay_vendor !!!
